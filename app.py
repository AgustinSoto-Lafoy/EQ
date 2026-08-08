import streamlit as st
import pandas as pd
import io
import math
import re
import numpy as np
from datetime import datetime
import logging

# =====================================
# CONFIGURACIÓN INICIAL
# =====================================

st.set_page_config(
    page_title="Cambio de Producto - Laminador",
    page_icon="⚙",  # solo la pestaña del navegador, no la interfaz
    layout="wide",
    initial_sidebar_state="collapsed"
)

# Configurar logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def _muestra_color(color: str, etiqueta: str) -> str:
    """Muestra del color EXACTO con que el Styler pinta la fila, junto a su etiqueta.

    Reemplaza los circulos emoji de las leyendas: su tono saturado no coincidia con el
    pastel real de las celdas, asi que la leyenda describia mal la tabla.
    """
    return (
        '<span style="display:inline-flex;align-items:center;gap:.5rem">'
        '<span style="width:.9rem;height:.9rem;border-radius:3px;flex:0 0 auto;'
        f'background:{color};border:1px solid rgba(0,0,0,.28)"></span>'
        f'<span>{etiqueta}</span></span>'
    )

# =====================================
# CAPACIDAD DE STANDS
# =====================================
# Posiciones del tren de laminación. `DU` (desbaste) queda fuera a propósito:
# no compite por los stands que el Equipo de Cambio monta y prepara.
POSICIONES_LINEA = ("M1", "M2", "M3", "M4", "A1", "A2", "A3", "A4", "A5", "A6")

# Valores iniciales del panel de capacidad; todos editables en la app.
TOTAL_STANDS_DEFAULT = 22
EN_MANTENCION_DEFAULT = 2
HORAS_PREP_STAND_DEFAULT = 2.0

# Dotación del taller: en cuántos stands se puede trabajar A LA VEZ. Es una
# restricción distinta de los stands disponibles (cuántos hay): tener 12 stands
# en taller no significa poder prepararlos los 12 simultáneamente.
PREPARACIONES_SIMULTANEAS_DEFAULT = 4

# Holgura bajo la cual un cambio se marca como ajustado en vez de holgado.
HOLGURA_AJUSTADA_H = 1.5

# =====================================
# NIVEL DE REGULACIÓN
# =====================================
# Una regulación no siempre cuesta lo mismo. Si solo cambia la calibración
# (Material, Luz) el stand se ajusta en línea; si además hay que cambiar una
# PIEZA de la guía, hay que ir a buscarla, desarmar y montar. Esa distinción es
# la que mueve la aguja en el tiempo de cambio, y es lo único que se busca aquí:
# separar los dos casos, no cuantificar cuánto más cuesta uno que otro.
#
# Criterio de operaciones (2026-08-06): "guía" son las piezas que hay que
# CAMBIAR. Lo que se ajusta sobre la guía ya montada no entra:
#   - `Diámetro Min - Max` es el rango de trabajo del producto, no una pieza
#     (ver la corrección del 2026-08-04: la tabla da lo que la guía ADMITE, el
#     DP el rango del producto, que debe caber dentro).
#   - `Estabilización` es la separación entre polines: se regula, no se cambia.
#
# `Canteo` SÍ es pieza, y esto corrige lo que decía `CONTEXTO_PROCESO.md` §6.
# No es el ángulo de torsión: es el polín canteador que mantiene la sección
# redonda en los extremos de una plana, y va montado en una guía EDG —
# "Edging Delivery Guide", literalmente salida canteadora— y a veces en un RTO.
PIEZAS_GUIA = frozenset({
    "Caja Guía Entrada", "Caja Guía Salida",
    "Embudo Entrada", "Embudo Salida",
    "Código Polín Entrada", "Código Polín Salida",
    "Canteo Entrada", "Canteo Salida",
    "Semiguía Entrada", "Semiguía Salida",
    "Raspador Entrada", "Raspador Salida",
    "Rodamiento Entrada", "Rodamiento Salida",
    "Ángulo Diagonal Entrada", "Ángulo Diagonal Salida",
})

# Etiquetas de nivel. Son centinelas: se comparan en `contar_regulaciones` y se
# muestran en tabla. Cambiar el texto obliga a cambiar ambos lados a la vez.
NIVEL_FUERTE = "Fuerte"
NIVEL_LEVE = "Leve"
NIVEL_NO_APLICA = "—"

# =====================================
# MODOS DE LECTURA DEL COMPARADOR
# =====================================
# La misma comparación la leen dos públicos con preguntas distintas:
#   - Taller  : qué componente cambia, para preparar el stand fuera de línea.
#               Es la vista histórica y el default; quien ya la usa no ve nada
#               distinto hasta que cambia el modo a propósito.
#   - Laminación: qué pasa en CADA STAND, para operar el cambio en línea.
#
# Son dos presentaciones de la MISMA clasificación, no dos cálculos. Ambas leen
# el detalle de `clasificar_cambios_codigo_canal`; si cada modo contara por su
# cuenta, el mismo cambio mostraría dos cifras según dónde se mire, que es
# exactamente la lección de `cb41c68`.
MODO_TALLER = "Taller"
MODO_LAMINACION = "Laminación"

# =====================================
# FUNCIONES DE CARGA DE DATOS
# =====================================

# Las hojas del Consolidado se ubican por CONTENIDO, no por posición.
# Leerlas por índice (`sheet_names[1]`) fallaba en silencio: bastaba insertar una
# hoja cualquiera para que se cargara la equivocada sin excepción ni error. El
# síntoma era engañoso — el aviso de Maestranza pasaba a listar los 71 códigos como
# "sin rendimiento registrado en el Consolidado" y ninguno estimaba cilindros —,
# culpando al dato cuando el dato estaba bien y el problema era qué hoja se leyó.


def _columnas_normalizadas(columnas):
    return [str(c).strip().lower() for c in columnas]


def _es_hoja_rendimientos(columnas):
    """Hoja de rendimientos: un código de canal y una columna de rendimiento.

    Usa los mismos criterios que `calcular_rango_rendimiento` para ubicar sus
    columnas, de modo que detector y consumidor no puedan divergir. El DDP tiene
    'Código Canal' pero no 'Rendimiento', así que no puede confundirse con esta.
    """
    cols = _columnas_normalizadas(columnas)
    tiene_codigo = any(c in ("código canal", "codigo canal") for c in cols)
    tiene_rendimiento = any("rendimiento" in c for c in cols)
    return tiene_codigo and tiene_rendimiento


def _es_hoja_ddp(columnas):
    """Hoja de diagramas de pase: 1 fila ≈ 1 pase (Producto × STD)."""
    cols = _columnas_normalizadas(columnas)
    tiene_codigo = any(c in ("código canal", "codigo canal") for c in cols)
    return "producto" in cols and "std" in cols and tiene_codigo


def _es_hoja_condiciones(columnas):
    """Condiciones de laminación del pie del formulario: Producto · Parámetro · Valor."""
    cols = _columnas_normalizadas(columnas)
    return "producto" in cols and any(c.startswith("par") and "metro" in c for c in cols) \
        and "valor" in cols


def _es_hoja_observaciones(columnas):
    """Observaciones numeradas al pie del Diagrama de Pase."""
    cols = _columnas_normalizadas(columnas)
    return "producto" in cols and "texto" in cols


def _es_hoja_versiones(columnas):
    """Metadatos del formulario: qué versión del DP es y de qué fecha."""
    cols = _columnas_normalizadas(columnas)
    tiene_version = any(c in ("versión", "version") for c in cols)
    tiene_fecha = any("fecha" in c for c in cols)
    return "producto" in cols and tiene_version and tiene_fecha


def _buscar_hoja(ruta, criterio):
    """Primera hoja de `ruta` cuyas columnas cumplen `criterio`.

    Devuelve `(nombre_hoja, df, hojas_disponibles)`; `(None, None, hojas)` si
    ninguna califica. No hay respaldo por posición a propósito: desactivar la
    funcionalidad con un aviso explícito es preferible a cargar la hoja
    equivocada y entregar resultados que parecen válidos.
    """
    xls = pd.ExcelFile(ruta)
    hojas = list(xls.sheet_names)
    for hoja in hojas:
        encabezado = pd.read_excel(xls, sheet_name=hoja, nrows=0)
        if criterio(encabezado.columns):
            return hoja, pd.read_excel(xls, sheet_name=hoja), hojas
    return None, None, hojas


@st.cache_data(ttl=3600)
def cargar_hojas_dp():
    """Hojas del pie del Diagrama de Pase: condiciones, observaciones y versiones.

    Va aparte de `cargar_datos` para no tocarle la firma: esa función devuelve 4
    valores y la leen varias pantallas. Agregarle tres más obligaría a cambiar
    todos los llamadores a la vez, que es el modo de falla de §6.14.

    Las tres son OPCIONALES por diseño. Un Consolidado que no las traiga —el
    vigente hasta hoy no las tenía— sigue funcionando: la vista lo dice y muestra
    lo que sí hay. Se ubican por CONTENIDO, nunca por nombre ni posición (§6.9).
    """
    hojas = {}
    for clave, criterio in (("condiciones", _es_hoja_condiciones),
                            ("observaciones", _es_hoja_observaciones),
                            ("versiones", _es_hoja_versiones)):
        try:
            nombre, df, _ = _buscar_hoja("data/Consolidado_Laminador.xlsx", criterio)
            hojas[clave] = df
            if df is not None:
                logger.info(f"Hoja de {clave} detectada: '{nombre}'")
        except Exception as e:
            logger.error(f"Error cargando hoja de {clave}: {str(e)}")
            hojas[clave] = None
    return hojas


@st.cache_data(ttl=3600)  # Cache por 1 hora
def cargar_datos():
    """Carga los archivos base de datos necesarios para la aplicación."""
    archivos = {
        "ddp": "data/Consolidado_Laminador.xlsx",
        "tiempo": "data/BBDD_Tiempo.xlsx", 
        "desbaste": "data/Diagrama_Desbaste.xlsx"
    }
    
    datos = {}
    errores = []
    
    for key, archivo in archivos.items():
        try:
            if key == "ddp":
                # El DDP también se ubica por contenido: una hoja insertada al
                # inicio haría que la app leyera cualquier cosa como diagrama de pases.
                hoja_ddp, df, hojas = _buscar_hoja(archivo, _es_hoja_ddp)
                if df is None:
                    raise ValueError(
                        "no se encontró la hoja de diagramas de pase (se esperaba una hoja con "
                        "columnas 'Producto', 'STD' y 'Código Canal'). "
                        f"Hojas encontradas: {', '.join(hojas)}"
                    )
                logger.info(f"Hoja de diagramas de pase detectada: '{hoja_ddp}'")
            else:
                df = pd.read_excel(archivo)
            # Optimizar tipos de datos básico
            # 'str' explícito: desde pandas 3 las columnas de texto son dtype 'str', e incluirlas
            # vía 'object' es retrocompatibilidad que pandas 4 retira. Sin esto, la conversión a
            # 'category' dejaría de aplicarse en silencio.
            for col in df.select_dtypes(include=['object', 'str']).columns:
                if df[col].nunique() / len(df) < 0.5:  # Solo si tiene sentido
                    try:
                        df[col] = df[col].astype('category')
                    except:
                        pass  # Si falla, mantener como object
            
            datos[key] = df
            logger.info(f"Archivo {archivo} cargado exitosamente")
        except Exception as e:
            errores.append(f"Error cargando {archivo}: {str(e)}")
            logger.error(f"Error cargando {archivo}: {str(e)}")

    # --- Rendimientos: hoja del Consolidado ubicada por contenido ---
    # Fuente adicional (no reemplaza la lectura anterior, que toma la hoja de
    # diagramas de pase). Si falla, no bloquea el resto de la app.
    df_rendimiento = None
    hojas_consolidado = []
    try:
        hoja_rend, df_rendimiento, hojas_consolidado = _buscar_hoja(
            archivos["ddp"], _es_hoja_rendimientos
        )
        if df_rendimiento is not None:
            logger.info(f"Hoja de rendimientos detectada: '{hoja_rend}'")
        else:
            logger.warning(
                "Ninguna hoja del Consolidado tiene columnas de código de canal y rendimiento. "
                f"Hojas encontradas: {', '.join(hojas_consolidado)}"
            )
    except Exception as e:
        logger.error(f"Error cargando hoja de rendimientos: {str(e)}")

    if errores:
        st.error("Errores al cargar archivos base:\n" + "\n".join(errores))
        return None, None, None, None

    if df_rendimiento is None:
        detalle = f" Hojas encontradas: {', '.join(hojas_consolidado)}." if hojas_consolidado else ""
        st.warning(
            "No se encontró la hoja de rendimientos dentro del Consolidado: ninguna hoja tiene "
            "a la vez una columna de código de canal y una de rendimiento." + detalle +
            " Los cálculos de canales/cilindros requeridos no estarán disponibles."
        )

    return datos.get("ddp"), datos.get("tiempo"), datos.get("desbaste"), df_rendimiento

def _cruzar_programa_con_mapa(archivo_bytes):
    """
    Lee el programa (.xlsm o .xlsx) y genera el DataFrame equivalente a TablaCombinada
    haciendo el cruce interno con la hoja 'Mapa'.

    Returns (df_resultado, warnings_list) o lanza Exception si falla algo crítico.
    """
    import openpyxl

    wb = openpyxl.load_workbook(archivo_bytes, read_only=True, keep_vba=True, data_only=True)

    # --- Validar hojas requeridas ---
    hojas_requeridas = ["Prog LAM REN", "Mapa"]
    for hoja in hojas_requeridas:
        if hoja not in wb.sheetnames:
            raise ValueError(
                f"El archivo no contiene la hoja requerida '{hoja}'. "
                f"Hojas disponibles: {wb.sheetnames}"
            )

    warnings = []

    # --- Leer Mapa ---
    # Contrato POSICIONAL: se leen las 4 primeras columnas y se les asignan estos
    # nombres, ignorando el encabezado real de la hoja.
    # `max_col=4` es obligatorio: el Mapa maestro trae 6 columnas (agrega "Recetas"
    # y "RM_LR"), y sin acotar, pd.DataFrame aborta la carga entera con
    # "4 columns passed, passed data had 6 columns".
    COLS_MAPA = ["Producto Limpio", "Nombre STD", "Producto STD", "Es Prueba"]
    ws_mapa = wb["Mapa"]

    # El contrato posicional falla en silencio si alguien reordena el Mapa: se leería
    # la columna equivocada sin error. Se valida el encabezado para que avise.
    encabezado = next(
        ws_mapa.iter_rows(min_row=1, max_row=1, max_col=4, values_only=True),
        (),
    )
    encabezado = [str(h).strip() if h is not None else "" for h in encabezado]
    if encabezado != COLS_MAPA:
        warnings.append(
            "Los encabezados A:D de la hoja 'Mapa' no son los esperados. "
            f"Se encontró {encabezado} y se esperaba {COLS_MAPA}. "
            "Se leyó igual por posición; verifica el orden de las columnas."
        )

    mapa_rows = list(ws_mapa.iter_rows(min_row=2, max_col=4, values_only=True))
    df_mapa = pd.DataFrame(
        mapa_rows,
        columns=COLS_MAPA
    ).dropna(subset=["Producto Limpio"])
    df_mapa["_key"] = df_mapa["Producto Limpio"].astype(str).str.strip()

    # --- Leer Prog LAM REN (encabezado en fila 8, datos desde fila 9) ---
    ws_prog = wb["Prog LAM REN"]
    all_rows = list(ws_prog.iter_rows(min_row=8, values_only=True))
    if not all_rows:
        raise ValueError("La hoja 'Prog LAM REN' no tiene datos.")

    headers = list(all_rows[0])
    # Renombrar PROGR. → PROGR para consistencia con el resto de la app
    headers = ["PROGR" if str(h) == "PROGR." else h for h in headers]

    data_rows = all_rows[1:]
    df_prog = pd.DataFrame(data_rows, columns=headers)

    # Cortar la secuencia en la primera fila con DESCRIPCIÓN vacía
    if df_prog["DESCRIPCIÓN"].isna().any():
        primera_vacia = df_prog["DESCRIPCIÓN"].isna().idxmax()
        df_prog = df_prog.iloc[:primera_vacia].copy()
    df_prog.reset_index(drop=True, inplace=True)

    if df_prog.empty:
        raise ValueError("No se encontraron filas con DESCRIPCIÓN en 'Prog LAM REN'.")

    # --- Cruce ---
    df_prog["_key"] = df_prog["DESCRIPCIÓN"].astype(str).str.strip()
    df_merged = df_prog.merge(
        df_mapa[["_key", "Nombre STD", "Producto STD", "Es Prueba"]].rename(
            columns={"Es Prueba": "Tabla13.Es Prueba"}
        ),
        on="_key",
        how="left"
    ).drop(columns=["_key"])

    # Calcular filas sin match para advertencia
    sin_match = df_merged[df_merged["Nombre STD"].isna()]["DESCRIPCIÓN"].dropna().unique().tolist()
    # `warnings` ya viene inicializada arriba: puede traer el aviso de encabezados.
    if sin_match:
        warnings.append(
            f"{len(sin_match)} producto(s) sin homologación en el Mapa: {', '.join(str(x) for x in sin_match[:5])}"
            + (" ..." if len(sin_match) > 5 else "")
        )

    return df_merged, warnings


def cargar_programa_usuario():
    """Maneja la carga del archivo de programa del usuario."""
    if "df_prog" not in st.session_state:
        with st.container():
            st.markdown("### Cargar Programa")
            archivo_programa = st.file_uploader(
                "Sube el archivo de programa (.xlsm o .xlsx)",
                type=None,  # Sin filtro MIME: Streamlit rechaza xlsm por su MIME type
                key="carga_global",
                help="Archivo debe contener las hojas 'Prog LAM REN' y 'Mapa'"
            )

            if archivo_programa is not None:
                ext = archivo_programa.name.rsplit(".", 1)[-1].lower()
                if ext not in ("xlsx", "xlsm"):
                    st.error(f"Formato no soportado (.{ext}). Sube un archivo .xlsx o .xlsm.")
                    archivo_programa = None

            if archivo_programa is not None:
                with st.spinner("Cargando y cruzando programa con Mapa..."):
                    try:
                        archivo_bytes = io.BytesIO(archivo_programa.read())
                        df_merged, warnings = _cruzar_programa_con_mapa(archivo_bytes)

                        st.session_state.df_prog = df_merged.reset_index(drop=True)

                        if warnings:
                            sin_match = df_merged[df_merged["Nombre STD"].isna()]["DESCRIPCIÓN"].dropna().unique().tolist()
                            with st.expander(f"{len(sin_match)} producto(s) sin homologación en el Mapa — requieren corrección", expanded=True):
                                st.warning("Estos productos fueron cargados pero **no tienen Nombre STD**. No podrán usarse en comparaciones hasta que se agreguen al Mapa.")
                                st.dataframe(
                                    pd.DataFrame({"DESCRIPCIÓN sin match en Mapa": sin_match}),
                                    width="stretch",
                                    hide_index=True
                                )

                        total = len(df_merged)
                        con_match = df_merged["Nombre STD"].notna().sum()
                        st.success(f"Programa cargado: {total} registros ({con_match} con homologación, {total - con_match} sin match)")
                        st.rerun()

                    except ValueError as e:
                        st.error(f"{e}")
                    except Exception as e:
                        st.error(f"Error al procesar archivo: {e}")
                        logger.error(f"Error cargando programa: {str(e)}")

# =====================================
# FUNCIONES DE COMPARACIÓN MEJORADAS
# =====================================

@st.cache_data
def comparar_productos(df_a, df_b, columnas):
    """Compara dos productos y retorna las diferencias - VERSIÓN SEGURA."""
    if df_a.empty or df_b.empty:
        return pd.DataFrame()
    
    resumen = []
    posiciones = sorted(set(df_a["STD"]).union(df_b["STD"]))
    
    for pos in posiciones:
        fila_a = df_a[df_a["STD"] == pos]
        fila_b = df_b[df_b["STD"] == pos]
        
        for col in columnas:
            # Verificar que la columna existe
            if col not in df_a.columns or col not in df_b.columns:
                continue
                
            val_a = fila_a[col].values[0] if not fila_a.empty and col in fila_a.columns else None
            val_b = fila_b[col].values[0] if not fila_b.empty and col in fila_b.columns else None
            
            # Saltar valores vacíos o nulos
            if (val_a is None or pd.isna(val_a)) and (val_b is None or pd.isna(val_b)):
                continue
            
            # Comparación segura
            try:
                cambia = val_a != val_b
            except (TypeError, ValueError):
                cambia = str(val_a) != str(val_b)
            
            resumen.append({
                "Posicion": pos,
                "Componente": col,
                "Valor A": str(val_a) if not pd.isna(val_a) else "-",
                "Valor B": str(val_b) if not pd.isna(val_b) else "-",
                "¿Cambia?": "Sí" if cambia else "No"
            })
    
    return pd.DataFrame(resumen)

@st.cache_data
def comparar_desbaste(df_desbaste, familia_a, familia_b):
    """Compara diagrama de desbaste entre dos familias"""
    try:
        # Si son la misma familia, no hay cambios
        if familia_a == familia_b:
            return pd.DataFrame()
        
        # Filtrar por familias de forma segura
        if familia_a != "(Todos)":
            desb_a = df_desbaste[df_desbaste["Familia"] == familia_a].copy()
        else:
            desb_a = df_desbaste.copy()
            
        if familia_b != "(Todos)":
            desb_b = df_desbaste[df_desbaste["Familia"] == familia_b].copy()
        else:
            desb_b = df_desbaste.copy()
        
        # Verificar columnas necesarias
        columnas_requeridas = ["SubSTD", "Componente limpio", "Valor"]
        if not all(col in df_desbaste.columns for col in columnas_requeridas):
            logger.error(f"Columnas faltantes en desbaste. Requeridas: {columnas_requeridas}")
            return pd.DataFrame()
        
        # Limpiar valores para comparación consistente
        def limpiar_valor(val):
            if pd.isna(val) or val is None:
                return None
            # Convertir a string y limpiar espacios
            val_str = str(val).strip()
            # Intentar convertir a número si es posible
            try:
                return float(val_str)
            except:
                return val_str
        
        # Aplicar limpieza a los valores
        desb_a["Valor_limpio"] = desb_a["Valor"].apply(limpiar_valor)
        desb_b["Valor_limpio"] = desb_b["Valor"].apply(limpiar_valor)
        
        # Obtener todos los pares únicos (SubSTD, Componente)
        pares_a = set(zip(desb_a["SubSTD"], desb_a["Componente limpio"]))
        pares_b = set(zip(desb_b["SubSTD"], desb_b["Componente limpio"]))
        todos_pares = sorted(pares_a.union(pares_b))
        
        resumen_desbaste = []
        for substd, comp in todos_pares:
            # Buscar valores de forma segura
            val_a_df = desb_a[(desb_a["SubSTD"] == substd) & (desb_a["Componente limpio"] == comp)]
            val_b_df = desb_b[(desb_b["SubSTD"] == substd) & (desb_b["Componente limpio"] == comp)]
            
            # Obtener valores limpios
            val_a_limpio = val_a_df["Valor_limpio"].iloc[0] if not val_a_df.empty else None
            val_b_limpio = val_b_df["Valor_limpio"].iloc[0] if not val_b_df.empty else None
            
            # Obtener valores originales para mostrar
            val_a_original = val_a_df["Valor"].iloc[0] if not val_a_df.empty else None
            val_b_original = val_b_df["Valor"].iloc[0] if not val_b_df.empty else None
            
            # Si ambos son None, continuar
            if val_a_limpio is None and val_b_limpio is None:
                continue
            
            # Determinar si hay cambio
            if val_a_limpio is None or val_b_limpio is None:
                # Si uno tiene valor y el otro no, es un cambio
                cambia = True
            else:
                # Comparar valores limpios
                cambia = val_a_limpio != val_b_limpio
                
            resumen_desbaste.append({
                "Posición": substd,
                "Componente": comp,
                "Valor A": str(val_a_original) if val_a_original is not None else "-",
                "Valor B": str(val_b_original) if val_b_original is not None else "-",
                "¿Cambia?": "Sí" if cambia else "No"
            })
        
        df_resultado = pd.DataFrame(resumen_desbaste)
        
        # Log para debugging
        if not df_resultado.empty:
            cambios = len(df_resultado[df_resultado["¿Cambia?"] == "Sí"])
            logger.info(f"Comparación desbaste {familia_a} vs {familia_b}: {cambios} cambios de {len(df_resultado)} componentes")
        
        return df_resultado
        
    except Exception as e:
        logger.error(f"Error en comparar_desbaste: {str(e)}")
        return pd.DataFrame()

def _parametros_cambiados(row_a, row_b, columnas):
    """Columnas cuyo valor difiere entre las dos filas.

    Dos nulos NO son un cambio: en este DDP el nulo es semántico ("no hay esa
    herramienta en el pase") y llega al 99% en algunas columnas.
    """
    cambiados = []
    for col in columnas:
        val_a = row_a.get(col)
        val_b = row_b.get(col)
        if (val_a is None or pd.isna(val_a)) and (val_b is None or pd.isna(val_b)):
            continue
        try:
            cambia = val_a != val_b
        except (TypeError, ValueError):
            cambia = str(val_a) != str(val_b)
        if cambia:
            cambiados.append(str(col))
    return cambiados


def _nivel_regulacion(parametros_cambiados):
    """Gradúa una regulación en Leve o Fuerte. Devuelve (nivel, motivo, piezas).

    Fuerte = hay que cambiar una PIEZA de guía, o sea buscarla, desarmar y
    montar. Leve = el stand se ajusta donde está. El motivo nombra las piezas
    primero porque son la razón del nivel; lo demás va detrás como contexto.

    `piezas` y `otros` se devuelven además del motivo para que el modo
    Laminación pueda listarlos sin parsear el texto. Es el mismo cálculo, no
    uno nuevo: un parser del motivo se desincronizaría en silencio en cuanto
    alguien cambiara la redacción.

    La tupla creció de 2 a 3 elementos y eso es compatible: los consumidores
    —incluida `verificar_nivel_regulacion.py`— indexan `[0]` y `[1]`.
    """
    piezas = [c for c in parametros_cambiados if c in PIEZAS_GUIA]
    otros = [c for c in parametros_cambiados if c not in PIEZAS_GUIA]

    if piezas:
        motivo = "cambia guía: " + ", ".join(piezas)
        if otros:
            motivo += " · además " + ", ".join(otros)
        return NIVEL_FUERTE, motivo, piezas

    if otros:
        return NIVEL_LEVE, "cambia: " + ", ".join(otros), []
    return NIVEL_LEVE, "sin cambios de parámetros", []


def contar_regulaciones(detalle):
    """Desglosa las regulaciones del detalle en (leves, fuertes).

    Existe como función única y no como cálculo repetido en cada pantalla: si
    el Comparador Manual y el Análisis de Secuencia contaran por su cuenta,
    podrían divergir y el mismo cambio mostraría dos cifras según dónde se mire
    (la lección de `cb41c68`).

    Cubre solo `POSICIONES_LINEA`, igual que los conteos de
    `clasificar_cambios_codigo_canal`: `DU` se muestra pero no se cuenta.
    """
    leves = fuertes = 0
    for d in detalle or []:
        if d.get("Posición") not in POSICIONES_LINEA:
            continue
        if d.get("Categoría") != "Regulación":
            continue
        if d.get("Nivel") == NIVEL_FUERTE:
            fuertes += 1
        else:
            leves += 1
    return leves, fuertes


@st.cache_data
def clasificar_cambios_codigo_canal(df_a, df_b):
    """
    Clasifica, posición por posición (STD), los cambios entre un producto
    origen (df_a) y un producto destino (df_b), distinguiendo dos motivos
    de "Regulación":

      A) El Código Canal cambia de posición pero sigue existiendo en el
         producto destino (existencia global, no solo por posición):
         "Regulación" — se reordena y calibra el mismo stand; NO corresponde
         cambio de stand.
      B) El Código Canal se mantiene igual en esa misma posición, pero algún
         otro parámetro técnico (Material, Luz, guías, embudos, ángulos,
         estabilización, etc.) cambia entre origen y destino: "Regulación" —
         el stand no se reemplaza ni se reubica, pero sí requiere ajuste/
         calibración por el cambio de condición.

    Solo cuando el Código Canal deja de existir por completo en el producto
    destino corresponde "Cambio completo" (reemplazar código de canal, guías,
    stand y elementos asociados).

    Ambos tipos de Regulación se gradúan además en un `Nivel` (ver `PIEZAS_GUIA`):
    "Leve" si solo hay que ajustar el stand en línea, "Fuerte" si además hay que
    cambiar una pieza de guía. El nivel se calcula igual en los dos casos —
    también en la reubicación, por criterio de operaciones del 2026-08-06: mover
    el código de posición no encarece por sí solo, lo que encarece es la pieza.

    Retorna (cambios_completos, regulaciones, detalle) donde detalle es una
    lista de dicts con Posición, Código Origen, Código Destino, Categoría, Nivel
    y Motivo, solo para las posiciones donde efectivamente hay una diferencia
    (de código o de algún otro parámetro técnico).

    La firma devuelve 3 valores a propósito: las suites de regresión hacen
    `_, _, det = ...`, y el desglose leve/fuerte se obtiene del detalle con
    `contar_regulaciones` en vez de agrandar la tupla.
    """
    detalle = []
    if (df_a.empty or df_b.empty
            or "Código Canal" not in df_a.columns or "Código Canal" not in df_b.columns
            or "STD" not in df_a.columns or "STD" not in df_b.columns):
        return 0, 0, detalle

    codigos_destino_existentes = set(df_b["Código Canal"].dropna())

    # Columnas de parámetros técnicos a revisar cuando el código NO cambia
    # (todo lo que no sea identificador de posición/producto ni el propio código)
    columnas_parametros = [
        col for col in df_a.columns
        if col not in ("STD", "Producto", "Familia", "Código Canal") and col in df_b.columns
    ]

    for _, row_a in df_a.iterrows():
        pos = row_a["STD"]
        codigo_a = row_a["Código Canal"]

        matching_b = df_b[df_b["STD"] == pos]
        if matching_b.empty:
            # Se mantiene el comportamiento previo: solo se evalúan posiciones
            # presentes en ambos productos.
            continue

        fila_b = matching_b.iloc[0]
        codigo_b = fila_b["Código Canal"]

        try:
            mismo_codigo = (codigo_a == codigo_b) or (pd.isna(codigo_a) and pd.isna(codigo_b))
        except (TypeError, ValueError):
            mismo_codigo = str(codigo_a) == str(codigo_b)

        parametros_cambiados = _parametros_cambiados(row_a, fila_b, columnas_parametros)

        if mismo_codigo:
            # El código no cambia: ¿cambia algún otro parámetro técnico en esta posición?
            if not parametros_cambiados:
                continue

            nivel, motivo, piezas = _nivel_regulacion(parametros_cambiados)
            detalle.append({
                "Posición": pos,
                "Código Origen": codigo_a if pd.notna(codigo_a) else "-",
                "Código Destino": codigo_b if pd.notna(codigo_b) else "-",
                "Categoría": "Regulación",
                "Nivel": nivel,
                "Motivo": "Mismo código; " + motivo,
                "Piezas": piezas,
                "Parámetros": parametros_cambiados,
            })
            continue

        # El código de canal difiere en esta posición: ¿el stand de origen sigue
        # existiendo en algún lugar del producto destino?
        if pd.notna(codigo_a) and codigo_a in codigos_destino_existentes:
            categoria = "Regulación"
            nivel, motivo_nivel, piezas = _nivel_regulacion(parametros_cambiados)
            motivo = "Código se reubica en otra posición"
            if parametros_cambiados:
                motivo += "; " + motivo_nivel
        else:
            categoria = "Cambio completo"
            # El cambio completo ya es la intervención más pesada: graduarlo
            # sería comparar contra sí mismo.
            nivel = NIVEL_NO_APLICA
            motivo = "Código deja de existir en el producto destino"
            piezas = [c for c in parametros_cambiados if c in PIEZAS_GUIA]

        detalle.append({
            "Posición": pos,
            "Código Origen": codigo_a if pd.notna(codigo_a) else "-",
            "Código Destino": codigo_b if pd.notna(codigo_b) else "-",
            "Categoría": categoria,
            "Nivel": nivel,
            "Motivo": motivo,
            "Piezas": piezas,
            "Parámetros": parametros_cambiados,
        })

    # Los conteos cubren solo las posiciones del tren (POSICIONES_LINEA). `DU`
    # sigue apareciendo en el detalle porque el operador necesita verlo, pero no
    # se cuenta: es desbaste y no compite por los stands del tren.
    en_linea = [d for d in detalle if d["Posición"] in POSICIONES_LINEA]
    cambios_completos = sum(1 for d in en_linea if d["Categoría"] == "Cambio completo")
    regulaciones = sum(1 for d in en_linea if d["Categoría"] == "Regulación")
    return cambios_completos, regulaciones, detalle


def stands_montados(df_ddp, producto):
    """Stands que este producto ocupa físicamente en el tren.

    No son siempre 10: un pase falso (`F`) deja la posición vacía, y ese stand
    queda libre para que el Equipo de Cambio lo prepare. Va de 6 a 10 según el
    producto. Devuelve None si el producto no está en el DDP.
    """
    if df_ddp is None or df_ddp.empty or "Producto" not in df_ddp.columns:
        return None
    g = df_ddp[(df_ddp["Producto"] == producto) & (df_ddp["STD"].isin(POSICIONES_LINEA))]
    if g.empty:
        return None
    reales = g[~g["Código Canal"].apply(_es_pase_falso)]
    return int(reales["STD"].nunique())


def stands_a_montar(detalle):
    """A partir del detalle de un cambio, separa montajes de liberaciones.

    Un "Cambio completo" cuyo código destino es `F` NO es un montaje: esa
    posición queda vacía en el producto destino, así que el stand se libera en
    vez de consumirse. Contarlo como montaje infla la necesidad de taller.
    """
    montar = liberar = 0
    for d in detalle or []:
        if d["Posición"] not in POSICIONES_LINEA or d["Categoría"] != "Cambio completo":
            continue
        if _es_pase_falso(d["Código Destino"]):
            liberar += 1
        else:
            montar += 1
    return montar, liberar


def horas_preparacion(n_montar, puestos, horas_por_stand):
    """Horas de taller para dejar listos `n_montar` stands.

    Se preparan en tandas del tamaño de los puestos disponibles: con 4 puestos y
    2 h por stand, 10 stands son 3 tandas = 6 h.
    """
    if not n_montar:
        return 0.0
    puestos = max(1, int(puestos or 1))
    return math.ceil(n_montar / puestos) * float(horas_por_stand)


def evaluar_capacidad_cambio(disponibles, montar, horas_bloque, puestos, horas_por_stand):
    """Decide si el cambio se puede dejar preparado. Devuelve (estado, motivo).

    Estados: "No evaluable" · "OK" · "Ajustado" · "Sin stands" · "Sin tiempo".
    `horas_bloque` es el tiempo de laminación del producto de origen, es decir
    la ventana real para preparar. Si el programa no trae horas, la restricción
    de tiempo no se evalúa y se dice explícitamente.
    """
    if disponibles is None or montar is None:
        return "No evaluable", "Falta homologación o datos de diagrama de pase"

    if montar == 0:
        return "OK", "Sin montajes: solo regulación en línea"

    if disponibles < montar:
        faltan = montar - disponibles
        return "Sin stands", f"Faltan {faltan} stand(s): se necesitan {montar} y hay {disponibles}"

    h_prep = horas_preparacion(montar, puestos, horas_por_stand)
    puestos = max(1, int(puestos or 1))
    tandas = math.ceil(montar / puestos)
    # Cómo se llega a las horas: hacerlo explícito evita que el número parezca salido de la nada.
    calculo = (f"{montar} stand(s) ÷ {puestos} simultáneo(s) = {tandas} tanda(s) × "
               f"{horas_por_stand:g} h = {h_prep:.1f} h")

    if horas_bloque is None or pd.isna(horas_bloque):
        return "OK", (f"{calculo}. Hay {disponibles} stand(s) disponibles. "
                      f"El programa no trae horas de bloque para contrastar el tiempo")

    if h_prep > horas_bloque:
        return "Sin tiempo", (f"{calculo}, pero el bloque dura {horas_bloque:.1f} h: "
                              f"faltan {h_prep - horas_bloque:.1f} h")

    holgura = horas_bloque - h_prep
    detalle = (f"{calculo} sobre un bloque de {horas_bloque:.1f} h; "
               f"sobran {holgura:.1f} h y {disponibles - montar} stand(s)")
    if holgura < HOLGURA_AJUSTADA_H:
        return "Ajustado", detalle
    return "OK", detalle


def horas_por_bloque(df_prog):
    """Horas de laminación de cada bloque, indexadas por el N° de grupo.

    Usa INICIO/FIN del programa. Devuelve None si el archivo no las trae — no
    todos los programas las tienen y la app no debe caerse por eso.
    """
    if df_prog is None or not {"INICIO", "FIN"}.issubset(df_prog.columns):
        return None
    try:
        d = df_prog.loc[:, ~df_prog.columns.duplicated()].copy()
        d["_ini"] = pd.to_datetime(d["INICIO"], errors="coerce")
        d["_fin"] = pd.to_datetime(d["FIN"], errors="coerce")
        if d["_ini"].isna().all() or d["_fin"].isna().all():
            return None
        d["_grupo"] = (d["Nombre STD"] != d["Nombre STD"].shift()).cumsum()
        bloques = d.groupby("_grupo").agg(ini=("_ini", "min"), fin=("_fin", "max"))
        horas = (bloques["fin"] - bloques["ini"]).dt.total_seconds() / 3600
        return horas.to_dict()
    except Exception as e:
        logger.error(f"Error calculando horas por bloque: {str(e)}")
        return None


@st.cache_data
def obtener_tiempo_cambio(df_tiempo, producto_origen, producto_destino):
    """
    Retorna el tiempo exacto para la dirección Origen → Destino.
    No hace fallback inverso: la dirección del cambio es relevante
    (124 pares asimétricos en BBDD_Tiempo).
    """
    try:
        columnas_requeridas = ["Nombre STD Origen", "Nombre STD Destino", "Minutos Cambio"]
        if not all(col in df_tiempo.columns for col in columnas_requeridas):
            return None
        mask = (
            (df_tiempo["Nombre STD Origen"] == producto_origen) &
            (df_tiempo["Nombre STD Destino"] == producto_destino)
        )
        resultado = df_tiempo.loc[mask, "Minutos Cambio"]
        return int(resultado.iloc[0]) if not resultado.empty else None
    except Exception as e:
        logger.error(f"Error obteniendo tiempo: {str(e)}")
        return None

def agrupar_cambios_consecutivos(df):
    """Agrupa cambios consecutivos del mismo tipo"""
    if df.empty:
        return df
    
    try:
        # Verificar columnas necesarias
        columnas_requeridas = ["Producto Origen", "Producto Destino"]
        if not all(col in df.columns for col in columnas_requeridas):
            return df
        
        # Crear grupos de forma segura
        df = df.copy()
        df["Grupo"] = (df[columnas_requeridas] != df[columnas_requeridas].shift()).any(axis=1).cumsum()

        # Definir agregaciones solo para columnas que existen
        agg_dict = {}
        columnas_posibles = {
            "Secuencia": "first",
            "Familia": "first", 
            "Producto Origen": "first",
            "Producto Destino": "first",
            "Tiempo estimado": "first",
            "Cambios Completo": "first",
            "Regulaciones": "first",
            # Nivel de regulación: mismo motivo que abajo, si no se listan se pierden
            "Reg. Fuertes": "first",
            "Reg. Leves": "first",
            # Capacidad de stands: si no se listan aquí, el groupby las descarta
            "Disponibles": "first",
            "A Montar": "first",
            "Libera": "first",
            "Horas Bloque": "first",
            "Estado": "first",
            "Motivo": "first"
        }
        
        for col, func in columnas_posibles.items():
            if col in df.columns:
                agg_dict[col] = func
        
        df_agrupado = df.groupby("Grupo", as_index=False).agg(agg_dict)
        return df_agrupado
        
    except Exception as e:
        logger.error(f"Error agrupando cambios: {str(e)}")
        return df

# =====================================
# FUNCIONES DE ESTILO Y UI
# =====================================

# Color de texto para las filas que Styler pinta con un pastel claro.
# Obligatorio: sin el, el texto hereda el color del tema y en modo oscuro queda
# claro sobre claro, ilegible. Los 13 fondos son claros, asi que va oscuro fijo.
TEXTO_PASTEL = "#16202B"


def resaltar_cambios(row):
    """Aplica estilo a las filas que tienen cambios."""
    try:
        color_cambio = "#ffebee"
        color_sin_cambio = "#f1f8e9"
        
        if "¿Cambia?" in row and row["¿Cambia?"] == "Sí":
            return [f'background-color: {color_cambio}; color: {TEXTO_PASTEL}; font-weight: bold'] * len(row)
        else:
            return [f'background-color: {color_sin_cambio}; color: {TEXTO_PASTEL}'] * len(row)
    except:
        return [''] * len(row)

def etiqueta_producto(nombre):
    """Cómo se MUESTRA y se BUSCA un producto en los selectores.

    9 de los 146 productos del DDP traen espacios internos dobles
    (`PLANA  50 x 6`, `CUADRADO  10`, `REDONDO 1 3/4"  (44.5) mm`), y la grafía
    es inconsistente dentro de una misma familia: `PLANA  50 x 6` lleva dos y
    `PLANA 50 x 4` uno. El filtro del `selectbox` compara subcadenas literales,
    así que escribir `PLANA 50 x 6` con un espacio no encuentra nada.

    Es un fallo especialmente malo de diagnosticar porque **un espacio de más es
    invisible en pantalla**: el producto no parece mal escrito, parece no existir.

    Se colapsa SOLO para mostrar; el selector sigue devolviendo el nombre
    original. Colapsarlo en el dato NO es opción: el cruce con el `Nombre STD`
    del Mapa es byte a byte y dejaría 26 filas sin diagrama de pase (§6.11).

    Seguro porque los 146 productos siguen siendo únicos al colapsar. Si alguna
    vez dos colapsaran al mismo texto, el selector mostraría dos opciones
    idénticas e indistinguibles; lo vigila `verificar_busqueda_productos.py`.
    """
    return re.sub(r"\s+", " ", str(nombre)).strip()


def mostrar_info_familia(producto, df_ddp, label):
    """Muestra información de la familia del producto."""
    try:
        if producto and "Familia" in df_ddp.columns:
            familia = df_ddp[df_ddp["Producto"] == producto]["Familia"].dropna().unique()
            if len(familia) > 0:
                st.info(f"{label} pertenece a la familia: **{familia[0]}**")
    except Exception as e:
        logger.error(f"Error mostrando familia: {str(e)}")

# =====================================
# DIAGRAMA DE PASE CON FORMATO
# =====================================
# Reproduce el formulario F-PGLAM101-06 con el que operaciones lee el DP en papel.
# El formato NO es cosmético: quien lo usa lleva años leyendo esa disposición, y una
# tabla con las mismas columnas en otro orden obliga a re-aprenderla.
#
# La estructura es de dos filas por stand — entrada `E` y salida `S` — con STD,
# código, material, luz, tolerancia y CTE combinados verticalmente sobre ambas.
# Por eso se arma en HTML y no con `st.dataframe`: hace falta `rowspan`, y un
# dataframe no combina celdas.

# Componentes del utilaje, en el orden del formulario impreso.
COMPONENTES_DP = [
    ("Caja Guía", "CAJA GUÍA"),
    ("Embudo", "EMBUDOS"),
    ("Código Polín", "POLÍN"),
    ("Diámetro Min - Max", "DIÁMETRO<br>MÁX - MÍN"),
    ("Ángulo Diagonal", "DIAGONAL<br>ÁNGULO"),
    ("Canteo", "CANTEO"),
    ("Estabilización", "ESTABILIZ"),
    ("Rodamiento", "RODAM."),
    ("Semiguía", "SEMIGUÍA"),
    ("Raspador", "RASPADOR"),
]

# Paleta del diagrama. Se fija `color` junto con `background-color` en cada regla:
# sin color de texto explícito, en modo oscuro queda claro sobre claro (§6.13).
DP_COLORES = {
    "encabezado": ("#1f4e79", "#ffffff"),   # franja de títulos
    "grupo": ("#dce6f1", "#1a1a1a"),        # sub-encabezados
    "std": ("#f2f2f2", "#1a1a1a"),          # columna de posición
    "entrada": ("#ffffff", "#1a1a1a"),
    "salida": ("#fafafa", "#1a1a1a"),
    "desbaste": ("#fff2cc", "#1a1a1a"),     # la fila DU, que no es del tren
}


def _txt(valor):
    """Celda a texto para mostrar. El nulo del utilaje es semántico —'no hay esa
    herramienta en el pase'— así que se muestra VACÍO, no como 'nan' ni '-'."""
    if valor is None:
        return ""
    try:
        if pd.isna(valor):
            return ""
    except (TypeError, ValueError):
        pass
    texto = str(valor).strip()
    return "" if texto.lower() in ("nan", "none", "nat") else texto


def _dp_filas_ordenadas(df_producto):
    """Filas del producto en la secuencia de la línea (DU → M1..M4 → A1..A6).

    No se rellenan los stands ausentes: no todos los productos recorren los 11 y
    inventar una fila vacía sería inventar un pase (§5).
    """
    orden = ["DU"] + list(POSICIONES_LINEA)
    filas = []
    for pos in orden:
        for _, fila in df_producto[df_producto["STD"] == pos].iterrows():
            filas.append(fila)
    # Cualquier STD que no esté en la secuencia canónica va al final, visible:
    # descartarlo en silencio escondería un dato que alguien escribió.
    conocidos = set(orden)
    for _, fila in df_producto[~df_producto["STD"].isin(conocidos)].iterrows():
        filas.append(fila)
    return filas


def construir_diagrama_pase(df_ddp, producto, hojas_dp=None):
    """Arma el diagrama de un producto desde el Consolidado y sus hojas de pie.

    Devuelve un dict con todo lo que necesitan la vista y el PDF, para que ambos
    salgan del MISMO armado: si cada salida consultara por su cuenta, la pantalla
    y el papel podrían mostrar cosas distintas del mismo producto.
    """
    df_producto = df_ddp[df_ddp["Producto"] == producto]
    filas = _dp_filas_ordenadas(df_producto)

    familia = ""
    if not df_producto.empty and "Familia" in df_producto.columns:
        familia = _txt(df_producto["Familia"].iloc[0])

    hojas_dp = hojas_dp or {}
    version, fecha_dp = "", ""
    df_ver = hojas_dp.get("versiones")
    if df_ver is not None and "Producto" in df_ver.columns:
        sub = df_ver[df_ver["Producto"] == producto]
        # Si hay varias versiones se muestra la vigente: es la que se está laminando.
        if "Vigente" in sub.columns and (sub["Vigente"].astype(str) == "Sí").any():
            sub = sub[sub["Vigente"].astype(str) == "Sí"]
        if not sub.empty:
            version = _txt(sub.iloc[0].get("Versión"))
            fecha_dp = _txt(sub.iloc[0].get("Fecha DP"))[:10]

    condiciones = []
    df_cond = hojas_dp.get("condiciones")
    if df_cond is not None and "Producto" in df_cond.columns:
        for _, r in df_cond[df_cond["Producto"] == producto].iterrows():
            condiciones.append((_txt(r.get("Grupo")), _txt(r.get("Parámetro")),
                                _txt(r.get("Valor"))))

    observaciones = []
    df_obs = hojas_dp.get("observaciones")
    if df_obs is not None and "Producto" in df_obs.columns:
        sub = df_obs[df_obs["Producto"] == producto]
        if "N°" in sub.columns:
            sub = sub.sort_values("N°")
        observaciones = [_txt(r.get("Texto")) for _, r in sub.iterrows()]

    return {
        "producto": producto,
        "familia": familia,
        "version": version,
        "fecha_dp": fecha_dp,
        "filas": filas,
        "condiciones": condiciones,
        "observaciones": observaciones,
    }


def _celda(contenido, fondo, texto, extra=""):
    return (f'<td style="background-color:{fondo};color:{texto};border:1px solid #999;'
            f'padding:2px 5px;font-size:11px;{extra}">{contenido}</td>')


def diagrama_pase_html(diagrama):
    """El diagrama como tabla HTML con el layout del formulario impreso."""
    fondo_enc, texto_enc = DP_COLORES["encabezado"]
    fondo_grp, texto_grp = DP_COLORES["grupo"]

    cols_pase = ["STD", "CÓDIGO<br>CANAL", "MATERIAL", "LUZ", "TOL.<br>+/-", "CTE.<br>%", "POS."]
    encabezado = "".join(
        f'<th style="background-color:{fondo_enc};color:{texto_enc};border:1px solid #999;'
        f'padding:3px 5px;font-size:10px;text-align:center;">{c}</th>'
        for c in cols_pase + [etiqueta for _, etiqueta in COMPONENTES_DP]
    )

    cuerpo = []
    for fila in diagrama["filas"]:
        std = _txt(fila.get("STD"))
        es_desbaste = std == "DU"
        clave_fondo = "desbaste" if es_desbaste else "std"
        fondo_std, texto_std = DP_COLORES[clave_fondo]

        # Las 6 primeras columnas son del PASE y se combinan sobre las dos filas
        # E/S. `rowspan` es un atributo del `<td>`, no una regla de estilo, así que
        # estas celdas no pueden salir de `_celda`.
        compartidas = "".join(
            f'<td rowspan="2" style="background-color:{fondo_std};color:{texto_std};'
            f'border:1px solid #999;padding:2px 5px;font-size:11px;text-align:center;'
            f'font-weight:{"600" if campo in ("STD", "Código Canal") else "400"};">'
            f'{_txt(fila.get(campo))}</td>'
            for campo in ("STD", "Código Canal", "Material", "Luz", "Tolerancia", "CTE %")
        )

        for posicion, etiqueta in (("Entrada", "E"), ("Salida", "S")):
            fondo, texto = DP_COLORES["desbaste" if es_desbaste else
                                      ("entrada" if posicion == "Entrada" else "salida")]
            celdas = [f'<td style="background-color:{fondo_grp};color:{texto_grp};'
                      f'border:1px solid #999;padding:2px 5px;font-size:11px;'
                      f'text-align:center;font-weight:600;">{etiqueta}</td>']
            for componente, _ in COMPONENTES_DP:
                celdas.append(_celda(_txt(fila.get(f"{componente} {posicion}")), fondo, texto))
            if posicion == "Entrada":
                cuerpo.append("<tr>" + compartidas + "".join(celdas) + "</tr>")
            else:
                cuerpo.append("<tr>" + "".join(celdas) + "</tr>")

    return (
        '<div style="overflow-x:auto;">'
        '<table style="border-collapse:collapse;width:100%;font-family:sans-serif;">'
        f"<thead><tr>{encabezado}</tr></thead>"
        f'<tbody>{"".join(cuerpo)}</tbody>'
        "</table></div>"
    )


def _latin1(texto):
    """Texto apto para fpdf 1.7.2, que escribe en latin-1 y no en UTF-8.

    Medido sobre todas las tablas del modelo: el único carácter fuera de latin-1
    es `→` (6 apariciones). Se traduce en vez de perderse. El `errors="replace"`
    final es la red para lo que aparezca mañana: un PDF con un carácter raro es
    molesto, uno que revienta al descargar deja al usuario sin nada.
    """
    if texto is None:
        return ""
    texto = str(texto)
    for origen, destino in (("→", "->"), ("←", "<-"), ("–", "-"), ("—", "-"),
                            ("“", '"'), ("”", '"'), ("’", "'")):
        texto = texto.replace(origen, destino)
    return texto.encode("latin-1", errors="replace").decode("latin-1")


def diagrama_pase_pdf(diagrama):
    """El diagrama como PDF horizontal a color. Devuelve bytes."""
    from fpdf import FPDF

    ANCHOS = [11, 20, 26, 12, 11, 10, 8]          # STD..POS
    ANCHOS += [20, 18, 18, 21, 17, 15, 15, 14, 20, 20]   # los 10 componentes

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=8)
    pdf.add_page()
    pdf.set_margins(6, 8, 6)

    # --- encabezado del formulario ---
    pdf.set_font("Helvetica", "B", 13)
    pdf.set_text_color(31, 78, 121)
    pdf.cell(0, 7, _latin1("DIAGRAMA DE PASES"), ln=1)
    pdf.set_font("Helvetica", "", 9)
    pdf.set_text_color(0, 0, 0)
    meta = f"FAMILIA: {diagrama['familia']}    PRODUCTO: {diagrama['producto']}"
    if diagrama["version"]:
        meta += f"    VERSIÓN DP: {diagrama['version']}"
    if diagrama["fecha_dp"]:
        meta += f"    FECHA: {diagrama['fecha_dp']}"
    pdf.cell(0, 5, _latin1(meta), ln=1)
    pdf.ln(1.5)

    # --- encabezado de la tabla ---
    titulos = ["STD", "CÓDIGO", "MATERIAL", "LUZ", "TOL.", "CTE%", "POS"]
    titulos += [e.replace("<br>", " ") for _, e in COMPONENTES_DP]
    pdf.set_font("Helvetica", "B", 6)
    pdf.set_fill_color(31, 78, 121)
    pdf.set_text_color(255, 255, 255)
    for ancho, titulo in zip(ANCHOS, titulos):
        pdf.cell(ancho, 6, _latin1(titulo), border=1, align="C", fill=True)
    pdf.ln()

    # --- cuerpo: dos filas por stand ---
    pdf.set_text_color(0, 0, 0)
    for fila in diagrama["filas"]:
        std = _txt(fila.get("STD"))
        es_desbaste = std == "DU"
        for posicion, etiqueta in (("Entrada", "E"), ("Salida", "S")):
            if es_desbaste:
                pdf.set_fill_color(255, 242, 204)      # el desbaste, distinguible
            elif posicion == "Entrada":
                pdf.set_fill_color(255, 255, 255)
            else:
                pdf.set_fill_color(245, 245, 245)
            # fpdf no combina celdas: el dato del pase se escribe en la fila de
            # entrada y la de salida lo deja en blanco. Es la convención del
            # formulario impreso, donde la celda combinada se lee igual.
            pdf.set_font("Helvetica", "B" if posicion == "Entrada" else "", 6)
            for ancho, campo in zip(ANCHOS[:6],
                                    ("STD", "Código Canal", "Material", "Luz",
                                     "Tolerancia", "CTE %")):
                valor = _txt(fila.get(campo)) if posicion == "Entrada" else ""
                pdf.cell(ancho, 5, _latin1(valor)[:16], border=1, align="C", fill=True)
            pdf.set_font("Helvetica", "B", 6)
            pdf.cell(ANCHOS[6], 5, etiqueta, border=1, align="C", fill=True)
            pdf.set_font("Helvetica", "", 6)
            for ancho, (componente, _) in zip(ANCHOS[7:], COMPONENTES_DP):
                valor = _txt(fila.get(f"{componente} {posicion}"))
                pdf.cell(ancho, 5, _latin1(valor)[:14], border=1, align="C", fill=True)
            pdf.ln()

    # --- condiciones de laminación ---
    if diagrama["condiciones"]:
        pdf.ln(2)
        pdf.set_font("Helvetica", "B", 8)
        pdf.set_text_color(31, 78, 121)
        pdf.cell(0, 5, _latin1("CONDICIONES DE LAMINACIÓN"), ln=1)
        pdf.set_text_color(0, 0, 0)
        pdf.set_font("Helvetica", "", 6.5)
        pdf.set_fill_color(220, 230, 241)
        for i, (grupo, parametro, valor) in enumerate(diagrama["condiciones"]):
            etiqueta = f"{grupo} · {parametro}" if grupo else parametro
            pdf.cell(70, 4.5, _latin1(etiqueta)[:52], border=1, fill=True)
            pdf.cell(40, 4.5, _latin1(valor)[:28], border=1)
            if i % 3 == 2:
                pdf.ln()
        if len(diagrama["condiciones"]) % 3:
            pdf.ln()

    # --- observaciones ---
    if diagrama["observaciones"]:
        pdf.ln(2)
        pdf.set_font("Helvetica", "B", 8)
        pdf.set_text_color(31, 78, 121)
        pdf.cell(0, 5, _latin1("OBSERVACIONES"), ln=1)
        pdf.set_text_color(0, 0, 0)
        pdf.set_font("Helvetica", "", 6.5)
        for texto in diagrama["observaciones"]:
            pdf.multi_cell(0, 4, _latin1(texto))

    salida = pdf.output(dest="S")
    # fpdf 1.7.2 devuelve `str` en Python 3; el PDF ya está en latin-1.
    return salida.encode("latin-1") if isinstance(salida, str) else bytes(salida)


def mostrar_diagrama_pase(df_ddp):
    """Pestaña: el Diagrama de Pase de un producto, con el formato del formulario."""
    st.markdown("### Diagrama de Pase")
    st.caption(
        "El diagrama se arma consultando el Consolidado y sus hojas de condiciones, "
        "observaciones y versiones. No es un archivo guardado: se construye al momento."
    )

    if df_ddp is None or df_ddp.empty or "Producto" not in df_ddp.columns:
        st.error("No hay diagramas de pase cargados.")
        return

    productos = sorted(df_ddp["Producto"].dropna().unique())
    # `format_func` es obligatorio: 9 de los 146 productos traen espacios dobles y
    # sin esto no se encuentran al escribir (§6.18).
    producto = st.selectbox("Producto", productos, key="dp_producto",
                            format_func=etiqueta_producto)
    if not producto:
        return

    hojas_dp = cargar_hojas_dp()
    diagrama = construir_diagrama_pase(df_ddp, producto, hojas_dp)

    faltantes = [n for n, c in (("condiciones", "condiciones"),
                                ("observaciones", "observaciones"),
                                ("versiones", "versiones")) if hojas_dp.get(c) is None]
    if faltantes:
        st.info(
            "El Consolidado cargado no trae hoja de " + ", ".join(faltantes) +
            ". El diagrama se muestra igual, con la tabla de pases."
        )

    cab = [f"**Familia:** {diagrama['familia'] or '—'}", f"**Producto:** {producto}"]
    if diagrama["version"]:
        cab.append(f"**Versión DP:** {diagrama['version']}")
    if diagrama["fecha_dp"]:
        cab.append(f"**Fecha:** {diagrama['fecha_dp']}")
    st.markdown(" · ".join(cab))

    st.markdown(diagrama_pase_html(diagrama), unsafe_allow_html=True)

    if diagrama["condiciones"]:
        st.markdown("#### Condiciones de laminación")
        st.dataframe(
            pd.DataFrame([{"Grupo": g, "Parámetro": p, "Valor": v}
                          for g, p, v in diagrama["condiciones"]]),
            width="stretch", hide_index=True
        )

    if diagrama["observaciones"]:
        st.markdown("#### Observaciones")
        for texto in diagrama["observaciones"]:
            st.markdown(f"- {texto}")

    try:
        pdf_bytes = diagrama_pase_pdf(diagrama)
        st.download_button(
            "Descargar diagrama en PDF",
            data=pdf_bytes,
            file_name=f"DP_{producto.replace('/', '-').replace(' ', '_')}.pdf",
            mime="application/pdf",
            key="dp_pdf",
        )
    except Exception as e:
        logger.error(f"Error generando el PDF del diagrama: {str(e)}")
        st.warning(f"No se pudo generar el PDF: {e}")


def mostrar_metricas_stands(cambios_completos, reg_fuertes, reg_leves, regulaciones):
    """Las 4 métricas de intervención por stand.

    Función única para que los dos modos muestren exactamente los mismos
    números: si cada vista armara sus métricas, podrían divergir.
    """
    col_r1, col_r2, col_r3, col_r4 = st.columns(4)
    with col_r1:
        st.metric("Cambios completos de stand", cambios_completos)
    with col_r2:
        st.metric("Regulaciones fuertes", reg_fuertes,
                  help="Además del ajuste hay que cambiar una pieza de guía: "
                       "caja guía, embudo, polín, canteo, semiguía, raspador, "
                       "rodamiento o ángulo diagonal.")
    with col_r3:
        st.metric("Regulaciones leves", reg_leves,
                  help="El stand se ajusta donde está: cambia la calibración "
                       "(material, luz) o un ajuste de la guía ya montada.")
    with col_r4:
        st.metric("Stands intervenidos", cambios_completos + regulaciones)


def _orden_stand(posicion):
    """Clave de orden para que los stands salgan en la secuencia de la línea.

    La secuencia canónica es DU → M1..M4 → A1..A6. Ordenar alfabéticamente
    pondría A1 antes que M1, o sea el final del tren antes del principio, que
    es justo al revés de como se recorre el cambio en planta.
    """
    if posicion == "DU":
        return -1
    try:
        return POSICIONES_LINEA.index(posicion)
    except ValueError:
        return len(POSICIONES_LINEA)


def _tabla_laminacion(detalle, categoria, nivel=None):
    """Filas del detalle de una categoría (y nivel), en orden de línea.

    Acota a `POSICIONES_LINEA` igual que `contar_regulaciones`, para que lo
    listado y lo contado nunca puedan diferir. `DU` se trata aparte.
    """
    filas = [
        d for d in detalle or []
        if d.get("Posición") in POSICIONES_LINEA
        and d.get("Categoría") == categoria
        and (nivel is None or d.get("Nivel") == nivel)
    ]
    return sorted(filas, key=lambda d: _orden_stand(d.get("Posición")))


def mostrar_resumen_laminacion(detalle):
    """Vista por stand del cambio, para quien lo opera en línea.

    Responde tres preguntas, en el orden en que importan en el laminador:
      1. ¿Qué stand se cambia completo, y de qué código a qué código?
      2. ¿Qué stand se regula fuerte, y qué guías hay que cambiar?
      3. ¿Qué stand se regula leve?

    Sale íntegramente del detalle de `clasificar_cambios_codigo_canal`: es
    presentación, no un cálculo nuevo. Los stands sin diferencias no aparecen
    —no hay nada que hacer en ellos— y por eso la suma de las tres tablas es
    igual a "Stands intervenidos" de la sección Resumen.
    """
    completos = _tabla_laminacion(detalle, "Cambio completo")
    fuertes = _tabla_laminacion(detalle, "Regulación", NIVEL_FUERTE)
    leves = _tabla_laminacion(detalle, "Regulación", NIVEL_LEVE)

    # --- 1. Cambio completo ---
    st.markdown(f"#### Cambio completo — {len(completos)} stand(s)")
    st.caption(
        "El código de canal deja de existir en el producto destino: hay que "
        "bajar el stand y montar otro ya preparado."
    )
    if completos:
        st.dataframe(
            pd.DataFrame([{
                "Stand": d["Posición"],
                "Código actual": str(d["Código Origen"]),
                "Código nuevo": str(d["Código Destino"]),
                # Un destino `F` no es un montaje: la posición queda vacía, así
                # que ese stand se LIBERA. Decirlo acá evita que se prepare un
                # stand que nadie va a montar (§6.10).
                "Observación": ("Pase falso: la posición queda vacía, el stand se libera"
                                if _es_pase_falso(d["Código Destino"]) else "Montar stand preparado"),
            } for d in completos]),
            width="stretch", hide_index=True
        )
    else:
        st.success("Ningún stand requiere cambio completo.")

    # --- 2. Regulación fuerte ---
    st.markdown(f"#### Regulación fuerte — {len(fuertes)} stand(s)")
    st.caption(
        "El stand se queda en línea, pero hay que cambiarle una pieza de guía: "
        "buscarla, desarmar y montar."
    )
    if fuertes:
        st.dataframe(
            pd.DataFrame([{
                "Stand": d["Posición"],
                "Código": str(d["Código Origen"]),
                "Guías a cambiar": ", ".join(d.get("Piezas") or []) or "—",
                "Además se ajusta": ", ".join(
                    [p for p in (d.get("Parámetros") or []) if p not in PIEZAS_GUIA]
                ) or "—",
            } for d in fuertes]),
            width="stretch", hide_index=True
        )
    else:
        st.success("Ningún stand requiere cambio de piezas de guía.")

    # --- 3. Regulación leve ---
    st.markdown(f"#### Regulación leve — {len(leves)} stand(s)")
    st.caption("El stand se ajusta donde está, sin cambiar piezas.")
    if leves:
        st.dataframe(
            pd.DataFrame([{
                "Stand": d["Posición"],
                "Código": str(d["Código Origen"]),
                "Qué se ajusta": ", ".join(d.get("Parámetros") or []) or "—",
            } for d in leves]),
            width="stretch", hide_index=True
        )
    else:
        st.success("Ningún stand requiere regulación leve.")

    # --- Desbaste, aparte y sin contarse ---
    # `DU` se muestra porque el operador necesita verlo, pero no entra en los
    # conteos del tren: es desbaste y no compite por los stands (§6.10).
    du = [d for d in detalle or [] if d.get("Posición") == "DU"]
    if du:
        with st.expander("Desbaste (DU) — se muestra, no se cuenta como stand del tren"):
            st.dataframe(
                pd.DataFrame([{
                    "Código actual": str(d["Código Origen"]),
                    "Código nuevo": str(d["Código Destino"]),
                    "Intervención": d["Categoría"],
                    "Motivo": d["Motivo"],
                } for d in du]),
                width="stretch", hide_index=True
            )


def mostrar_metricas_resumen(df_cambios):
    """Muestra métricas de resumen de cambios."""
    try:
        if not df_cambios.empty and "¿Cambia?" in df_cambios.columns:
            total_cambios = len(df_cambios[df_cambios["¿Cambia?"] == "Sí"])
            total_componentes = len(df_cambios)

            # Sin "% Cambios": la proporción de componentes que cambian no dice
            # cuánto trabajo implica el cambio. Eso lo responden las métricas de
            # stands (completos / regulaciones) de la sección Resumen.
            col1, col2 = st.columns(2)
            with col1:
                st.metric("Total Cambios", total_cambios)
            with col2:
                st.metric("Total Componentes", total_componentes)
    except Exception as e:
        logger.error(f"Error mostrando métricas: {str(e)}")

# =====================================
# INTERFAZ PRINCIPAL
# =====================================

def main():
    """Función principal de la aplicación."""
    
    # Header
    st.markdown("""
    # Plataforma de Cambio de Producto – Laminador Renca
    *Sistema de análisis y comparación de productos para optimización de cambios*
    """)
    
    # Cargar datos base
    with st.spinner("Cargando datos base..."):
        df_ddp, df_tiempo, df_desbaste, df_rendimiento = cargar_datos()
    
    if df_ddp is None:
        st.error("No se pudieron cargar los datos base. Verifica que los archivos existan.")
        st.stop()
    
    # Mostrar información básica
    with st.expander("Información de datos cargados", expanded=False):
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("Productos DDP", len(df_ddp))
        with col2:
            st.metric("Tiempos de Cambio", len(df_tiempo))
        with col3:
            st.metric("Registros Desbaste", len(df_desbaste))
    
    # Carga del programa
    cargar_programa_usuario()
    
    # Pestañas principales
    tabs = st.tabs([
        "Diagrama de Pase",
        "Comparador Manual",
        "Análisis de Secuencia",
        "Resumen Maestranza",
        "Utilaje"
    ])

    # PESTAÑA 1: DIAGRAMA DE PASE
    # Va primera porque es la consulta más básica —"cómo se lamina este producto"—
    # y no depende de tener un programa cargado.
    with tabs[0]:
        mostrar_diagrama_pase(df_ddp)

    # PESTAÑA 2: COMPARADOR MANUAL
    with tabs[1]:
        st.subheader("Comparación Manual de Productos")
        mostrar_comparador_manual(df_ddp, df_tiempo, df_desbaste)
    
    # PESTAÑA 3: SECUENCIA DE PROGRAMA
    with tabs[2]:
        st.subheader("Análisis de Secuencia de Programa")
        if "df_prog" in st.session_state:
            mostrar_secuencia_programa(df_ddp, df_tiempo)
        else:
            st.info("Por favor carga primero el archivo de programa.")

    # PESTAÑA 4: MAESTRANZA
    with tabs[3]:
        st.subheader("Resumen Técnico para Maestranza")
        if "df_prog" in st.session_state:
            mostrar_resumen_maestranza(df_ddp, df_rendimiento)
        else:
            st.info("Por favor carga primero el archivo de programa.")

    # PESTAÑA 5: UTILAJE
    with tabs[4]:
        st.subheader("Análisis de Utilaje")
        mostrar_analisis_utilaje(df_ddp)

def mostrar_comparador_manual(df_ddp, df_tiempo, df_desbaste):
    """Muestra el comparador manual de productos."""
    
    # Verificar columnas necesarias
    if "Familia" not in df_ddp.columns or "Producto" not in df_ddp.columns:
        st.error("El archivo DDP debe contener las columnas 'Familia' y 'Producto'")
        return
    
    # Selección de familias con ancho uniforme
    familias = ["(Todos)"] + sorted(df_ddp["Familia"].dropna().unique())
    
    col_f1, col_f2 = st.columns([2, 2])
    
    with col_f1:
        familia_a = st.selectbox("Familia A", familias, key="famA")
    with col_f2:
        familia_b = st.selectbox("Familia B", familias, key="famB")
    
    # Filtrar productos por familia
    try:
        if familia_a == "(Todos)":
            df_fam_a = df_ddp
        else:
            df_fam_a = df_ddp[df_ddp["Familia"] == familia_a]
            
        if familia_b == "(Todos)":
            df_fam_b = df_ddp
        else:
            df_fam_b = df_ddp[df_ddp["Familia"] == familia_b]
        
        productos_a = sorted(df_fam_a["Producto"].dropna().unique())
        productos_b = sorted(df_fam_b["Producto"].dropna().unique())
        
    except Exception as e:
        st.error(f"Error filtrando productos: {str(e)}")
        return
    
    # Selección de productos con ancho uniforme
    col_a, col_b = st.columns([2, 2])
    
    with col_a:
        if productos_a:
            producto_a = st.selectbox("Producto A", productos_a, key="A",
                                      format_func=etiqueta_producto)
            if familia_a == "(Todos)":
                mostrar_info_familia(producto_a, df_ddp, "Producto A")
        else:
            st.warning("No hay productos disponibles para la Familia A")
            return
    
    with col_b:
        if productos_b:
            producto_b = st.selectbox("Producto B", productos_b, key="B",
                                      format_func=etiqueta_producto)
            if familia_b == "(Todos)":
                mostrar_info_familia(producto_b, df_ddp, "Producto B")
        else:
            st.warning("No hay productos disponibles para la Familia B")
            return
    
    # Mostrar comparación
    if producto_a and producto_b:
        if producto_a == producto_b:
            st.warning("Has seleccionado el mismo producto en ambos lados.")
        else:
            mostrar_comparacion_productos(
                df_ddp, df_tiempo, df_desbaste, 
                producto_a, producto_b, familia_a, familia_b
            )

def mostrar_comparacion_productos(df_ddp, df_tiempo, df_desbaste, producto_a, producto_b, familia_a, familia_b):
    try:
        df_a = df_ddp[df_ddp["Producto"] == producto_a]
        df_b = df_ddp[df_ddp["Producto"] == producto_b]
        
        if df_a.empty or df_b.empty:
            st.warning("No se encontraron datos para uno o ambos productos.")
            return
        
        # Tiempo de cambio — respetar direccionalidad (A→B puede ≠ B→A)
        tiempo_ab = obtener_tiempo_cambio(df_tiempo, producto_a, producto_b)
        tiempo_ba = obtener_tiempo_cambio(df_tiempo, producto_b, producto_a)

        if tiempo_ab is None and tiempo_ba is None:
            st.warning("**Tiempo de cambio:** No registrado para estos productos")
        elif tiempo_ab == tiempo_ba or tiempo_ba is None:
            st.success(f"**Tiempo de cambio:** {tiempo_ab} min")
        elif tiempo_ab is None:
            st.success(f"**Tiempo de cambio:** {tiempo_ba} min (solo dirección inversa registrada)")
        else:
            # Asimétrico: mostrar ambas direcciones
            col_t1, col_t2 = st.columns(2)
            with col_t1:
                st.success(f"**{producto_a} → {producto_b}:** {tiempo_ab} min")
            with col_t2:
                st.info(f"**{producto_b} → {producto_a}:** {tiempo_ba} min")
        
        # Comparación técnica (DDP)
        st.markdown("---")
        st.markdown("### Análisis Técnico")

        # La misma comparación, dos lecturas. Taller es el default a propósito:
        # es la vista histórica, así que quien ya usa la app no ve nada distinto
        # hasta que cambia el modo deliberadamente.
        modo_analisis = st.segmented_control(
            "Modo de lectura",
            options=[MODO_TALLER, MODO_LAMINACION],
            default=MODO_TALLER,
            required=True,          # sin esto, un clic sobre la opción activa la deselecciona
            key="modo_analisis",
            help="**Taller**: el detalle componente por componente, para preparar el stand. "
                 "**Laminación**: qué hay que hacer en cada stand, agrupado por tipo de "
                 "intervención.",
        )

        # Se clasifica UNA sola vez y la usan los dos modos. Que ambos lean el
        # mismo detalle es lo que garantiza que no puedan mostrar cifras
        # distintas del mismo cambio.
        try:
            cambios_completos, regulaciones, detalle_reg = clasificar_cambios_codigo_canal(df_a, df_b)
            reg_leves, reg_fuertes = contar_regulaciones(detalle_reg)
        except Exception as e:
            logger.error(f"Error clasificando cambios de stand: {str(e)}")
            cambios_completos, regulaciones = 0, 0
            reg_leves, reg_fuertes = 0, 0
            detalle_reg = []

        if modo_analisis == MODO_LAMINACION:
            mostrar_metricas_stands(cambios_completos, reg_fuertes, reg_leves, regulaciones)
            st.markdown("---")
            mostrar_resumen_laminacion(detalle_reg)
            return

        # ---- De aquí en adelante, modo Taller (vista original) ----

        # Opción de filtro encima de las tablas
        col_filtro, col_space = st.columns([1, 3])
        with col_filtro:
            mostrar_solo_cambios = st.checkbox("Solo mostrar cambios", value=True, key="filtro_tablas")

        columnas_ddp = [col for col in df_a.columns if col not in ["STD", "Producto", "Familia"]]

        # Se inicializa antes del `if`: la sección Resumen la lee más abajo y sin
        # esto, un producto sin columnas técnicas lanzaba NameError.
        resumen_ddp = pd.DataFrame()

        if columnas_ddp:
            with st.spinner("Analizando diferencias técnicas..."):
                resumen_ddp = comparar_productos(df_a, df_b, columnas_ddp)
            
            if not resumen_ddp.empty:
                mostrar_metricas_resumen(resumen_ddp)
                
                if mostrar_solo_cambios:
                    resumen_filtrado = resumen_ddp[resumen_ddp["¿Cambia?"] == "Sí"]
                    if resumen_filtrado.empty:
                        st.success("**¡No hay cambios técnicos entre estos productos!**")
                    else:
                        st.dataframe(
                            resumen_filtrado.style.apply(resaltar_cambios, axis=1),
                            width="stretch"
                        )
                else:
                    st.dataframe(
                        resumen_ddp.style.apply(resaltar_cambios, axis=1),
                        width="stretch"
                    )
        
        # Comparación desbaste
        st.markdown("---")
        st.markdown("### Análisis de Diagrama Desbaste")
        
        # Obtener las familias de los productos seleccionados
        familia_real_a = df_a["Familia"].iloc[0] if not df_a.empty and "Familia" in df_a.columns else familia_a
        familia_real_b = df_b["Familia"].iloc[0] if not df_b.empty and "Familia" in df_b.columns else familia_b
        
        # Log para debugging
        logger.info(f"Comparando desbaste: Producto A '{producto_a}' (Familia {familia_real_a}) vs Producto B '{producto_b}' (Familia {familia_real_b})")
        
        with st.spinner("Analizando diagrama de desbaste..."):
            df_desbaste_cmp = comparar_desbaste(df_desbaste, familia_real_a, familia_real_b)
        
        if not df_desbaste_cmp.empty:
            mostrar_metricas_resumen(df_desbaste_cmp)
            
            if mostrar_solo_cambios:
                desbaste_filtrado = df_desbaste_cmp[df_desbaste_cmp["¿Cambia?"] == "Sí"]
                if desbaste_filtrado.empty:
                    st.success("**¡No hay cambios en el diagrama de desbaste!**")
                else:
                    st.dataframe(
                        desbaste_filtrado.style.apply(resaltar_cambios, axis=1),
                        width="stretch"
                    )
            else:
                st.dataframe(
                    df_desbaste_cmp.style.apply(resaltar_cambios, axis=1),
                    width="stretch"
                )
        else:
            st.info("No se encontraron datos de desbaste para comparar.")
        
        # ===============================
        # NUEVO RESUMEN: Cambios por Componente
        # ===============================
        st.markdown("---")
        st.markdown("### Resumen")
        st.caption(
            "Solo el tren laminador (M1–M4, A1–A6). El desbaste no se cuenta: ni la "
            "posición DU del diagrama de pase, ni el diagrama de desbaste, que se "
            "compara por familia y es otra unidad de análisis."
        )

        # Intervención por stand. Se clasificó una sola vez al entrar al Análisis
        # Técnico y ese resultado se reusa acá: `clasificar_cambios_codigo_canal`
        # ya restringe sus conteos a POSICIONES_LINEA (§6.10), así que DU queda
        # fuera por construcción.
        mostrar_metricas_stands(cambios_completos, reg_fuertes, reg_leves, regulaciones)

        # La tabla por componente sale SOLO del diagrama de pase y SOLO de las
        # posiciones del tren. `df_desbaste_cmp` ya no se concatena.
        resumen_total = resumen_ddp

        if not resumen_total.empty and "¿Cambia?" in resumen_total.columns:
            solo_tren = resumen_total[resumen_total["Posicion"].isin(POSICIONES_LINEA)]
            resumen_componentes = (
                solo_tren[solo_tren["¿Cambia?"] == "Sí"]
                .groupby("Componente")
                .size()
                .reset_index(name="Cantidad de Cambios")
                .sort_values("Cantidad de Cambios", ascending=False)
            )

            if not resumen_componentes.empty:
                st.dataframe(resumen_componentes, width="stretch", hide_index=True)
            else:
                st.success("No se registraron cambios en ningún componente del tren.")
        else:
            st.info("No se encontraron diferencias para construir el resumen de componentes.")

            
    except Exception as e:
        st.error(f"Error en la comparación: {str(e)}")
        logger.error(f"Error en mostrar_comparacion_productos: {str(e)}")

def mostrar_secuencia_programa(df_ddp, df_tiempo):
    """Muestra el análisis de la secuencia del programa."""
    
    try:
        df_prog = st.session_state.df_prog
        st.markdown(f"**Programa cargado:** {len(df_prog)} registros")

        # --- Capacidad de taller (editable) ---
        with st.expander("Capacidad de stands", expanded=False):
            st.markdown(
                "El análisis cruza **dos restricciones distintas**:\n\n"
                "- **¿Tengo los stands?** Los disponibles no son un número fijo: se calculan como "
                "`total − mantención − los que ocupa el producto en línea`. Un producto con pases "
                "falsos ocupa menos de 10 posiciones, así que **libera** stands para montaje "
                "(van de 10 a 14 según el producto).\n"
                "- **¿Alcanza el tiempo?** Depende de la **dotación del taller** —en cuántos stands "
                "se puede trabajar a la vez— y de las horas del bloque que se está laminando, que "
                "es la ventana real para preparar el cambio siguiente."
            )
            c1, c2, c3, c4 = st.columns(4)
            with c1:
                total_stands = st.number_input(
                    "Total de stands", min_value=1, max_value=99,
                    value=TOTAL_STANDS_DEFAULT, step=1,
                    help="Parque completo de stands de la planta."
                )
            with c2:
                en_mantencion = st.number_input(
                    "En mantención", min_value=0, max_value=99,
                    value=EN_MANTENCION_DEFAULT, step=1,
                    help="Fuera de servicio. Varía; ajústalo a la situación real."
                )
            with c3:
                puestos_taller = st.number_input(
                    "Preparaciones simultáneas", min_value=1, max_value=20,
                    value=PREPARACIONES_SIMULTANEAS_DEFAULT, step=1,
                    help="Dotación del taller: en cuántos stands puede trabajar el equipo AL MISMO "
                         "TIEMPO. No es lo mismo que los stands disponibles (cuántos hay): con 12 "
                         "stands en taller y dotación para 4, se preparan en 3 tandas."
                )
            with c4:
                horas_prep = st.number_input(
                    "Horas por stand", min_value=0.5, max_value=48.0,
                    value=HORAS_PREP_STAND_DEFAULT, step=0.5,
                    help="Desmontar, cambiar canal y guías, calibrar."
                )

        # Horas de laminación de cada bloque: la ventana real para preparar.
        horas_bloques = horas_por_bloque(df_prog)
        if horas_bloques is None:
            st.info(
                "El programa no trae columnas `INICIO`/`FIN`, así que no se puede evaluar "
                "si alcanza el **tiempo** para preparar. Se evalúa solo la disponibilidad de stands."
            )
        grupos_prog = (df_prog["Nombre STD"] != df_prog["Nombre STD"].shift()).cumsum()

        with st.spinner("Analizando secuencia de cambios..."):
            resumen = []

            for i in range(len(df_prog) - 1):
                origen = df_prog.loc[i, "Nombre STD"]
                destino = df_prog.loc[i + 1, "Nombre STD"]

                # Skip si es el mismo producto
                if origen == destino:
                    continue

                # Obtener tiempo de cambio
                tiempo = obtener_tiempo_cambio(df_tiempo, origen, destino)
                
                # Clasificar cambios de código canal: Cambio completo vs Regulación
                # (un stand solo se cuenta como "cambio completo" si deja de existir
                # en el producto destino; si se reubica en otra posición, es Regulación)
                cambios_completos = 0
                regulaciones = 0
                reg_leves = reg_fuertes = 0
                montar = liberar = None
                try:
                    df_a = df_ddp[df_ddp["Producto"] == origen]
                    df_b = df_ddp[df_ddp["Producto"] == destino]

                    if not df_a.empty and not df_b.empty and "Código Canal" in df_a.columns:
                        cambios_completos, regulaciones, detalle_cap = clasificar_cambios_codigo_canal(df_a, df_b)
                        montar, liberar = stands_a_montar(detalle_cap)
                        reg_leves, reg_fuertes = contar_regulaciones(detalle_cap)
                except Exception as e:
                    logger.error(f"Error calculando cambios código canal: {str(e)}")

                # --- Capacidad: ¿se puede dejar montado este cambio? ---
                try:
                    ocupados = stands_montados(df_ddp, origen)
                    disponibles = None if ocupados is None else int(total_stands) - int(en_mantencion) - ocupados
                    h_bloque = horas_bloques.get(grupos_prog.iloc[i]) if horas_bloques else None
                    estado, motivo = evaluar_capacidad_cambio(
                        disponibles, montar, h_bloque, puestos_taller, horas_prep
                    )
                except Exception as e:
                    logger.error(f"Error evaluando capacidad de stands: {str(e)}")
                    disponibles, h_bloque = None, None
                    estado, motivo = "No evaluable", "Error al evaluar la capacidad"


                # Obtener familias
                try:
                    familia_origen = df_ddp[df_ddp["Producto"] == origen]['Familia'].iloc[0] if not df_ddp[df_ddp["Producto"] == origen].empty else "N/A"
                    familia_destino = df_ddp[df_ddp["Producto"] == destino]['Familia'].iloc[0] if not df_ddp[df_ddp["Producto"] == destino].empty else "N/A"
                except:
                    familia_origen = "N/A"
                    familia_destino = "N/A"
                
                resumen.append({
                    "Secuencia": i + 1,
                    "Familia": f"{familia_origen} → {familia_destino}",
                    "Producto Origen": origen,
                    "Producto Destino": destino,
                    "Tiempo estimado": tiempo,
                    "Cambios Completo": cambios_completos,
                    "Regulaciones": regulaciones,
                    "Reg. Fuertes": reg_fuertes,
                    "Reg. Leves": reg_leves,
                    "Disponibles": disponibles,
                    "A Montar": montar,
                    "Libera": liberar,
                    "Horas Bloque": h_bloque,
                    "Estado": estado,
                    "Motivo": motivo
                })
        
        if not resumen:
            st.info("No se encontraron cambios de producto en la secuencia.")
            return
        
        # Agrupar cambios consecutivos
        df_resumen = agrupar_cambios_consecutivos(pd.DataFrame(resumen))
        
        # Mostrar resumen general
        try:
            tiempo_total = df_resumen["Tiempo estimado"].dropna().sum() if "Tiempo estimado" in df_resumen.columns else 0
            cambios_totales = len(df_resumen)
            
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("Total de Cambios", cambios_totales)
            with col2:
                st.metric("Tiempo Total Estimado", f"{tiempo_total:.0f} min" if tiempo_total > 0 else "N/A")
            with col3:
                st.metric("Tiempo Promedio/Cambio", f"{tiempo_total/cambios_totales:.1f} min" if tiempo_total > 0 and cambios_totales > 0 else "N/A")
        except Exception as e:
            logger.error(f"Error mostrando métricas secuencia: {str(e)}")

        # --- Capacidad de stands a lo largo del programa ---
        try:
            if "Estado" in df_resumen.columns:
                st.markdown("---")
                st.markdown("### Capacidad de Stands")

                conteo = df_resumen["Estado"].value_counts()
                criticos = df_resumen[df_resumen["Estado"].isin(["Sin stands", "Sin tiempo"])]
                ajustados = df_resumen[df_resumen["Estado"] == "Ajustado"]
                no_eval = df_resumen[df_resumen["Estado"] == "No evaluable"]

                # Horizonte: cuántos cambios seguidos alcanzan a dejarse montados
                # con los stands libres al inicio, antes de reponer.
                horizonte = 0
                acumulado = 0
                disp_inicial = None
                for _, fila in df_resumen.iterrows():
                    if pd.isna(fila.get("Disponibles")) or pd.isna(fila.get("A Montar")):
                        break
                    if disp_inicial is None:
                        disp_inicial = int(fila["Disponibles"])
                    acumulado += int(fila["A Montar"])
                    if acumulado > disp_inicial:
                        break
                    horizonte += 1

                c1, c2, c3 = st.columns(3)
                with c1:
                    if disp_inicial is not None:
                        st.metric("Horizonte de montaje", f"{horizonte} cambio(s)",
                                  help=f"Cuántos cambios seguidos alcanzas a dejar montados "
                                       f"con los {disp_inicial} stands libres al inicio, sin reponer.")
                    else:
                        st.metric("Horizonte de montaje", "N/A")
                with c2:
                    st.metric("Cambios críticos", len(criticos),
                              help="Sin stands suficientes o sin tiempo para prepararlos.")
                with c3:
                    st.metric("Margen ajustado", len(ajustados),
                              help=f"Alcanzan por menos de {HOLGURA_AJUSTADA_H:.1f} h.")

                if not criticos.empty:
                    lineas = [
                        f"- **Cambio #{int(f['Secuencia'])}** · {f['Producto Origen']} → "
                        f"{f['Producto Destino']} — {f['Estado']}: {f['Motivo']}"
                        for _, f in criticos.iterrows()
                    ]
                    st.error("**Cambios que no se pueden dejar preparados:**\n" + "\n".join(lineas))

                if not ajustados.empty:
                    lineas = [
                        f"- **Cambio #{int(f['Secuencia'])}** · {f['Producto Origen']} → "
                        f"{f['Producto Destino']} — {f['Motivo']}"
                        for _, f in ajustados.iterrows()
                    ]
                    st.warning("**Margen ajustado (alcanza, pero sin holgura):**\n" + "\n".join(lineas))

                if not no_eval.empty:
                    st.info(
                        f"{len(no_eval)} cambio(s) no evaluable(s) por falta de homologación en el "
                        "Mapa o de diagrama de pase. Un producto sin homologar ciega también los "
                        "cambios vecinos: corregirlo en el Mapa maestro recupera esa visibilidad."
                    )

                columnas_cap = [c for c in ["Secuencia", "Producto Origen", "Producto Destino",
                                            "Disponibles", "A Montar", "Libera", "Horas Bloque",
                                            "Estado", "Motivo"] if c in df_resumen.columns]
                tabla_cap = df_resumen[columnas_cap].copy()
                if "Horas Bloque" in tabla_cap.columns:
                    tabla_cap["Horas Bloque"] = tabla_cap["Horas Bloque"].map(
                        lambda v: f"{v:.1f}" if pd.notna(v) else "—"
                    )
                st.dataframe(tabla_cap, width="stretch", hide_index=True)
        except Exception as e:
            logger.error(f"Error mostrando capacidad de stands: {str(e)}")
            st.warning("No se pudo calcular la capacidad de stands.")

        # Mostrar cambios detallados
        st.markdown("---")
        st.markdown("### Detalle de Cambios en Secuencia")
        
        for idx, fila in df_resumen.iterrows():
            try:
                tiempo_mostrar = f"{int(fila['Tiempo estimado'])} min" if pd.notna(fila.get('Tiempo estimado')) else "No registrado"
                
                # Color coding para el tiempo
                if pd.notna(fila.get('Tiempo estimado')):
                    if fila['Tiempo estimado'] > 60:
                        tiempo_color = "🔴"
                    elif fila['Tiempo estimado'] > 30:
                        tiempo_color = "🟡"
                    else:
                        tiempo_color = "🟢"
                else:
                    tiempo_color = "⚪"
                
                cambios_completos_fila = fila.get('Cambios Completo', 0)
                regulaciones_fila = fila.get('Regulaciones', 0)
                fuertes_fila = fila.get('Reg. Fuertes', 0)
                leves_fila = fila.get('Reg. Leves', 0)
                secuencia = fila.get('Secuencia', idx + 1)
                origen = fila.get('Producto Origen', 'N/A')
                destino = fila.get('Producto Destino', 'N/A')
                
                estado_fila = fila.get("Estado")
                icono_estado = {
                    "OK": "🟢", "Ajustado": "🟡",
                    "Sin stands": "🔴", "Sin tiempo": "🔴", "No evaluable": "⚪"
                }.get(estado_fila, "")
                sufijo_estado = f" | {icono_estado} {estado_fila}" if estado_fila else ""

                # El desglose va en el título y no dentro del expander: quien
                # revisa la secuencia decide cuál abrir sin abrirlos todos.
                if pd.notna(fuertes_fila) and (fuertes_fila or leves_fila):
                    detalle_reg = f"{int(fuertes_fila)} fuerte(s) / {int(leves_fila)} leve(s)"
                else:
                    detalle_reg = f"{regulaciones_fila} regulación(es)"

                titulo = f"{tiempo_color} **Cambio #{secuencia}** | {origen} → {destino} | {tiempo_mostrar} | {cambios_completos_fila} cambio(s) completo(s) | {detalle_reg}{sufijo_estado}"

                with st.expander(titulo):
                    if estado_fila:
                        montar_fila = fila.get("A Montar")
                        libera_fila = fila.get("Libera")
                        disp_fila = fila.get("Disponibles")
                        resumen_cap = f"**{estado_fila}** — {fila.get('Motivo', '')}"
                        if estado_fila in ("Sin stands", "Sin tiempo"):
                            st.error(resumen_cap)
                        elif estado_fila == "Ajustado":
                            st.warning(resumen_cap)
                        elif estado_fila == "No evaluable":
                            st.info(resumen_cap)
                        else:
                            st.success(resumen_cap)

                        if pd.notna(montar_fila):
                            m1, m2, m3 = st.columns(3)
                            with m1:
                                st.metric("Stands disponibles",
                                          int(disp_fila) if pd.notna(disp_fila) else "—")
                            with m2:
                                st.metric("A montar", int(montar_fila))
                            with m3:
                                st.metric("Libera (pase falso)",
                                          int(libera_fila) if pd.notna(libera_fila) else 0)

                    df_a_cmp = df_ddp[df_ddp["Producto"] == origen]
                    df_b_cmp = df_ddp[df_ddp["Producto"] == destino]
                    
                    if not df_a_cmp.empty and not df_b_cmp.empty:
                        if "Código Canal" in df_a_cmp.columns:
                            _, _, detalle_stand = clasificar_cambios_codigo_canal(df_a_cmp, df_b_cmp)
                            if detalle_stand:
                                st.markdown("**Clasificación de cambios de stand (Código Canal):**")
                                st.dataframe(pd.DataFrame(detalle_stand), width="stretch", hide_index=True)
                                st.caption(
                                    f"**Nivel** — `{NIVEL_FUERTE}`: hay que cambiar una pieza de guía "
                                    "(caja guía, embudo, polín, canteo, semiguía, raspador, rodamiento "
                                    f"o ángulo diagonal). `{NIVEL_LEVE}`: el stand se ajusta donde está. "
                                    f"`{NIVEL_NO_APLICA}`: no aplica, el cambio completo ya es la "
                                    "intervención más pesada."
                                )
                                if any(d["Posición"] not in POSICIONES_LINEA for d in detalle_stand):
                                    st.caption(
                                        "`DU` (desbaste) se muestra pero no se cuenta: no compite por "
                                        "los stands del tren. Un destino `F` es pase falso — esa posición "
                                        "queda vacía, así que libera un stand en vez de consumirlo."
                                    )

                        columnas_cmp = [col for col in df_a_cmp.columns if col not in ["STD", "Producto", "Familia"]]
                        resumen_cmp = comparar_productos(df_a_cmp, df_b_cmp, columnas_cmp)
                        
                        if not resumen_cmp.empty:
                            resumen_cmp_cambios = resumen_cmp[resumen_cmp["¿Cambia?"] == "Sí"]
                            
                            if not resumen_cmp_cambios.empty:
                                st.dataframe(resumen_cmp_cambios, width="stretch")
                            else:
                                st.success("No hay cambios técnicos para este cambio de producto")
                        else:
                            st.info("No se pudieron analizar las diferencias técnicas")
                    else:
                        st.warning("No se encontraron datos para uno o ambos productos")
                        
            except Exception as e:
                logger.error(f"Error mostrando cambio {idx}: {str(e)}")
                st.error(f"Error mostrando cambio {idx + 1}")
        
    except Exception as e:
        st.error(f"Error analizando secuencia: {str(e)}")
        logger.error(f"Error en mostrar_secuencia_programa: {str(e)}")

def _exportar_maestranza_xlsx(df_resumen, df_frecuencia, df_prog_completo):
    """
    Genera el Excel de Maestranza con Table style (TableStyleLight15),
    columna Nota en Frecuencia_Cilindros, y nombre de columna Toneladas.
    Retorna un BytesIO listo para st.download_button.
    """
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.utils import get_column_letter

    def _escribir_tabla(ws, df, table_name, table_style="TableStyleLight15"):
        """Escribe df en ws como Excel Table con estilo."""
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        n_rows = len(df) + 1  # +1 header
        n_cols = len(df.columns)
        ref = f"A1:{get_column_letter(n_cols)}{n_rows}"
        tab = Table(displayName=table_name, ref=ref)
        style = TableStyleInfo(
            name=table_style,
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        tab.tableStyleInfo = style
        ws.add_table(tab)
        # Autofit columnas (estimado por contenido)
        for col_cells in ws.columns:
            max_len = max(
                (len(str(c.value)) if c.value is not None else 0) for c in col_cells
            )
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 4, 40)

    wb = Workbook()

    # --- Hoja 1: Resumen_Maestranza ---
    ws1 = wb.active
    ws1.title = "Resumen_Maestranza"
    _escribir_tabla(ws1, df_resumen, "Tabla1")

    # --- Hoja 2: Frecuencia_Cilindros ---
    if not df_frecuencia.empty:
        ws2 = wb.create_sheet("Frecuencia_Cilindros")
        # Insertar columna Nota vacía en posición 2 (después de Código Canal)
        df_frec_export = df_frecuencia.copy().rename(columns={"Toneladas": "Toneladas"})
        df_frec_export.insert(1, "Nota", None)
        _escribir_tabla(ws2, df_frec_export, "Tabla2")

    # --- Hoja 3: Programa_Completo ---
    ws3 = wb.create_sheet("Programa_Completo")
    for r in dataframe_to_rows(df_prog_completo, index=False, header=True):
        ws3.append(r)
    # Sin Table style en programa (igual al archivo deseado)
    for col_cells in ws3.columns:
        max_len = max(
            (len(str(c.value)) if c.value is not None else 0) for c in col_cells
        )
        ws3.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 4, 50)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# Sufijos que indican moleteado, ordenados de más largo a más corto para que la
# sustitución no corte una variante por la mitad. El DDP escribe 'CP 23 Mol',
# 'PL Molet.' u 'OP 4209 MLTDO'; la hoja de rendimientos escribe siempre 'CP 23 M'.
_SUFIJOS_MOLETEADO = ("MOLETEADO", "MOLET", "MLTDO", "MOL")

# 'F' = pase falso: el pase va vacío. No es un código de canal — no hay geometría,
# no consume canales ni cilindros — por lo que queda fuera del cálculo de
# canales/cilindros requeridos. Sí se muestra en el diagrama de pases, donde
# significa justamente que esa posición no se usa.
#
# Los Diagramas de Pase vigentes lo anotan CON la posición (`F A5`, `F M3`, `FA 6`),
# que tras normalizar queda `FA5`/`FM3`/`FA6`. Se aceptan todas: la posición ya la da
# la fila del diagrama, así que dentro del código es redundante (criterio de
# operaciones, 2026-08-03). Reconocerlas no es cosmético — un pase falso tomado por
# canal real vuelve a contarse como cilindro (§6.8) y como stand a montar (§6.10).
#
# El patrón queda ACOTADO a las posiciones del tren en vez de aceptar cualquier
# código que empiece con F: si algún día aparece un canal real tipo `FR 12`, debe
# seguir contándose como canal. Hoy no existe ninguno — medido sobre el DDP y sobre
# la hoja de rendimientos, lo único que empieza con F es el pase falso.
_CODIGOS_PASE_FALSO = frozenset(
    {"F"} | {f"F{posicion}" for posicion in ("DU",) + POSICIONES_LINEA}
)


def _normalizar_codigo_canal(valor):
    """
    Normaliza un código de canal para cruzarlo entre archivos con distinta
    escritura (espacios, guiones, comas, mayúsculas/minúsculas). Se usa
    únicamente como clave interna de cruce; en pantalla siempre se muestra
    el código tal como aparece en el DDP/Programa.

    El sufijo de moleteado se unifica a 'M'. La distinción sí importa: 'CP 23' y
    'CP 23 M' son canales distintos —el segundo lleva moleteado, un proceso
    posterior al mecanizado que trabaja la superficie del canal para mejorar el
    agarre—, por lo que solo se unifica la GRAFÍA del sufijo; nunca se elimina.
    """
    if pd.isna(valor):
        return None
    s = str(valor).strip().upper()
    for ch in (" ", "-", ".", ","):
        s = s.replace(ch, "")
    for sufijo in _SUFIJOS_MOLETEADO:
        if s.endswith(sufijo):
            return s[:-len(sufijo)] + "M"
    return s


def _es_pase_falso(valor):
    """True si el código corresponde a un pase falso (pase vacío, sin canal)."""
    return _normalizar_codigo_canal(valor) in _CODIGOS_PASE_FALSO


@st.cache_data
def calcular_rango_rendimiento(df_rendimiento):
    """
    A partir de la segunda hoja del Consolidado (Código Canal, Dureza, Calidad,
    Rendimiento [t/canal], N° Canales), construye por código de canal el
    rendimiento mínimo y máximo disponible.

    Cuando un código de canal tiene más de una calidad, la hoja de rendimientos
    entrega dos valores (rendimiento inferior y superior). Ambos se conservan;
    nunca se elige uno arbitrariamente. El N° de Canales queda emparejado con
    el rendimiento de su propia fila (misma calidad), para que el cálculo de
    cilindros sea consistente.
    """
    columnas_salida = [
        "_codigo_norm",
        "Rendimiento Mín [t/canal]", "Rendimiento Máx [t/canal]",
        "N Canales (Rend Mín)", "N Canales (Rend Máx)"
    ]
    if df_rendimiento is None or df_rendimiento.empty:
        return pd.DataFrame(columns=columnas_salida)

    df = df_rendimiento.copy()

    col_codigo = next((c for c in df.columns if str(c).strip().lower() in ("código canal", "codigo canal")), None)
    col_rend = next((c for c in df.columns if "rendimiento" in str(c).strip().lower()), None)
    col_ncanales = next(
        (c for c in df.columns
         if "canales" in str(c).strip().lower()
         and "código" not in str(c).strip().lower()
         and "codigo" not in str(c).strip().lower()),
        None
    )

    if col_codigo is None or col_rend is None:
        logger.error("La hoja de rendimientos no tiene las columnas esperadas (Código Canal / Rendimiento).")
        return pd.DataFrame(columns=columnas_salida)

    df["_codigo_norm"] = df[col_codigo].apply(_normalizar_codigo_canal)

    # El rendimiento y el N° de canales deben ser numéricos: más abajo se dividen.
    # La hoja trae celdas con anotaciones de texto (p. ej. '600 (200 en A&B)') que
    # no se pueden interpretar sin ambigüedad. Se descartan en vez de adivinar un
    # valor: el código queda como "sin rendimiento" y aparece en el aviso de la
    # pestaña Maestranza, para que se corrija en la fuente.
    df[col_rend] = pd.to_numeric(df[col_rend], errors="coerce")
    if col_ncanales:
        df[col_ncanales] = pd.to_numeric(df[col_ncanales], errors="coerce")

    df = df.dropna(subset=["_codigo_norm", col_rend])

    filas = []
    for codigo_norm, grupo in df.groupby("_codigo_norm"):
        fila_min = grupo.loc[grupo[col_rend].idxmin()]
        fila_max = grupo.loc[grupo[col_rend].idxmax()]
        filas.append({
            "_codigo_norm": codigo_norm,
            "Rendimiento Mín [t/canal]": fila_min[col_rend],
            "Rendimiento Máx [t/canal]": fila_max[col_rend],
            "N Canales (Rend Mín)": fila_min[col_ncanales] if col_ncanales else None,
            "N Canales (Rend Máx)": fila_max[col_ncanales] if col_ncanales else None,
        })

    return pd.DataFrame(filas, columns=columnas_salida)


def _calcular_canales_cilindros(frecuencia_en_programa, df_rendimiento):
    """
    Usa las toneladas ya calculadas por código de canal (frecuencia_en_programa)
    y el rendimiento (mín/máx) de la segunda hoja del Consolidado para estimar:
      - Canales Requeridos (mín / máx)
      - Cilindros Requeridos (mín / máx)

    Cuando el código tiene dos calidades (rango de rendimiento), se conserva el
    rango completo: no se elige un valor arbitrario. Los códigos sin rendimiento
    registrado quedan con estas columnas en blanco y marcados en 'Observación'.
    """
    df_rango = calcular_rango_rendimiento(df_rendimiento)

    df = frecuencia_en_programa.copy()
    df["_codigo_norm"] = df["Código Canal"].apply(_normalizar_codigo_canal)
    df = df.merge(df_rango, on="_codigo_norm", how="left")

    def _canales(toneladas, rendimiento):
        if pd.isna(rendimiento) or rendimiento == 0:
            return None
        return math.ceil(toneladas / rendimiento)

    df["Canales Requeridos Mín"] = df.apply(
        lambda r: _canales(r["Toneladas"], r["Rendimiento Máx [t/canal]"]), axis=1
    )
    df["Canales Requeridos Máx"] = df.apply(
        lambda r: _canales(r["Toneladas"], r["Rendimiento Mín [t/canal]"]), axis=1
    )

    def _cilindros(row):
        opciones = []
        c_min = row.get("Canales Requeridos Mín")
        c_max = row.get("Canales Requeridos Máx")
        ncan_de_rend_max = row.get("N Canales (Rend Máx)")   # emparejado con Canales Requeridos Mín
        ncan_de_rend_min = row.get("N Canales (Rend Mín)")   # emparejado con Canales Requeridos Máx

        if c_min is not None and pd.notna(ncan_de_rend_max) and ncan_de_rend_max:
            opciones.append(math.ceil(c_min / ncan_de_rend_max))
        if c_max is not None and pd.notna(ncan_de_rend_min) and ncan_de_rend_min:
            opciones.append(math.ceil(c_max / ncan_de_rend_min))

        if not opciones:
            return pd.Series([None, None])
        return pd.Series([min(opciones), max(opciones)])

    df[["Cilindros Requeridos Mín", "Cilindros Requeridos Máx"]] = df.apply(_cilindros, axis=1)

    df["Observación"] = df["Rendimiento Mín [t/canal]"].apply(
        lambda v: "Sin rendimiento registrado en Consolidado" if pd.isna(v) else ""
    )

    df = df.drop(columns=["_codigo_norm", "N Canales (Rend Mín)", "N Canales (Rend Máx)"])
    return df


def mostrar_resumen_maestranza(df_ddp, df_rendimiento=None):
    """Muestra el resumen técnico para maestranza con análisis de cilindros."""
    
    try:
        df_prog = st.session_state.df_prog.copy()
        
        with st.spinner("Generando resumen para maestranza..."):
            # Detectar bloques consecutivos del mismo producto
            df_prog["Grupo"] = (df_prog["Nombre STD"] != df_prog["Nombre STD"].shift()).cumsum()

            # Verificar que existe la columna PROGR
            if "PROGR" not in df_prog.columns:
                st.error("El archivo de programa debe contener la columna 'PROGR' para calcular toneladas")
                return

            # --- AUDITORÍA 1: filas con PROGR nulo ---
            filas_sin_progr = df_prog[df_prog["PROGR"].isna()]
            productos_sin_progr = filas_sin_progr["Nombre STD"].dropna().unique().tolist()

            # Agrupar y sumar toneladas (excluir PROGR nulo para evitar sum→0)
            df_programa = (
                df_prog.dropna(subset=["PROGR"])
                .groupby(["Grupo", "Nombre STD"], as_index=False)
                .agg({"PROGR": "sum"})
                .rename(columns={"PROGR": "Toneladas"})
            )
            df_programa = df_programa[df_programa["Toneladas"] > 0].reset_index(drop=True)
            df_programa["Toneladas"] = df_programa["Toneladas"].astype(int)

            # Productos con PROGR nulo que NO aparecen con tonelaje real en ningún otro bloque
            productos_en_programa = set(df_programa["Nombre STD"].unique())
            productos_solo_nulos = [p for p in productos_sin_progr if p not in productos_en_programa]

            # Seleccionar primeras ocurrencias por Producto y STD para posiciones específicas
            posiciones_deseadas = ["M1", "M2", "M3", "M4", "A1", "A2", "A3", "A4", "A5", "A6"]

            # Verificar que tenemos las columnas necesarias
            if "STD" in df_ddp.columns and "Código Canal" in df_ddp.columns:
                df_canal_unico = (
                    df_ddp[df_ddp["STD"].isin(posiciones_deseadas)]
                    .dropna(subset=["Código Canal"])
                    .sort_values(["Producto", "STD"])
                    .drop_duplicates(subset=["Producto", "STD"], keep="first")
                )

                if not df_canal_unico.empty:
                    df_canal_pivot = df_canal_unico.pivot(
                        index="Producto",
                        columns="STD",
                        values="Código Canal"
                    ).reset_index()
                    df_canal_pivot.columns.name = None

                    # --- AUDITORÍA 2: productos del programa sin match en DDP ---
                    productos_en_ddp = set(df_canal_pivot["Producto"].unique())
                    productos_sin_ddp = [
                        p for p in df_programa["Nombre STD"].unique()
                        if p not in productos_en_ddp
                    ]

                    df_resumen = df_programa.merge(
                        df_canal_pivot,
                        left_on="Nombre STD",
                        right_on="Producto",
                        how="left"
                    ).drop(columns=["Producto"], errors='ignore')

                    # --- AUDITORÍA 3: posiciones faltantes por producto ---
                    posiciones_presentes = [p for p in posiciones_deseadas if p in df_resumen.columns]
                    audit_posiciones = []
                    for _, row in df_resumen.iterrows():
                        faltantes = [
                            pos for pos in posiciones_presentes
                            if pd.isna(row.get(pos))
                        ]
                        if faltantes:
                            audit_posiciones.append({
                                "Producto": row["Nombre STD"],
                                "Posiciones sin datos": ", ".join(faltantes)
                            })

                    columnas_orden = ["Nombre STD", "Toneladas"] + posiciones_deseadas
                    df_resumen = df_resumen[[col for col in columnas_orden if col in df_resumen.columns]]
                else:
                    df_resumen = df_programa[["Nombre STD", "Toneladas"]]
                    productos_sin_ddp = list(df_programa["Nombre STD"].unique())
                    audit_posiciones = []
            else:
                df_resumen = df_programa[["Nombre STD", "Toneladas"]]
                productos_sin_ddp = list(df_programa["Nombre STD"].unique())
                audit_posiciones = []
                st.warning("No se encontraron columnas 'STD' o 'Código Canal' para análisis detallado")
        
        # Mostrar métricas generales
        try:
            total_toneladas = df_resumen["Toneladas"].sum()
            productos_unicos = df_resumen["Nombre STD"].nunique()
            
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("Total Toneladas", f"{total_toneladas:,.0f}")
            with col2:
                st.metric("Productos Únicos", productos_unicos)
            with col3:
                bloques_consecutivos = len(df_resumen)
                st.metric("Bloques Consecutivos", bloques_consecutivos)
        except Exception as e:
            logger.error(f"Error calculando métricas maestranza: {str(e)}")
        
        # -----------------------------------------------
        # AUDITORÍAS — sección consolidada
        # -----------------------------------------------
        hay_alertas = bool(productos_solo_nulos or productos_sin_ddp or audit_posiciones)

        if hay_alertas:
            n_criticos = len(productos_sin_ddp)
            n_advertencias = len(productos_solo_nulos) + len(audit_posiciones)
            titulo = f"Inconsistencias detectadas — {n_criticos} crítica(s), {n_advertencias} advertencia(s)"

            with st.expander(titulo, expanded=True):

                if productos_sin_ddp:
                    st.error(f"**{len(productos_sin_ddp)} producto(s) no encontrados en DDP** — columnas M1–A6 vacías y sin datos en frecuencia de cilindros. Verifica que el nombre STD coincida exactamente con el Consolidado_Laminador.")
                    st.dataframe(
                        pd.DataFrame({"Producto sin datos en DDP": productos_sin_ddp}),
                        width="stretch", hide_index=True
                    )

                if productos_solo_nulos:
                    if productos_sin_ddp:
                        st.markdown("---")
                    st.warning(f"**{len(productos_solo_nulos)} producto(s) excluidos por PROGR vacío** — aparecen en el programa solo con tonelaje en blanco. Si deben producirse, corrige el programa.")
                    st.dataframe(
                        pd.DataFrame({"Producto excluido (PROGR vacío)": productos_solo_nulos}),
                        width="stretch", hide_index=True
                    )

                if audit_posiciones:
                    if productos_sin_ddp or productos_solo_nulos:
                        st.markdown("---")
                    st.warning(f"**{len(audit_posiciones)} producto(s) con posiciones M1–A6 sin Código Canal** — puede ser intencional si la posición no se usa, o un dato faltante en el DDP.")
                    st.dataframe(
                        pd.DataFrame(audit_posiciones),
                        width="stretch", hide_index=True
                    )
        else:
            st.success("Sin inconsistencias — todos los productos tienen datos completos.")

        # Tabla principal
        st.markdown("### Resumen Detallado por Producto")
        st.dataframe(df_resumen, width="stretch")
        
        # ===============================================
        # SECCIÓN DE FRECUENCIA DE CILINDROS
        # ===============================================
        
        st.markdown("### Frecuencia de Cilindros en Programa")
        
        try:
            # Crear una lista con todos los códigos de canal para cada producto en el programa
            codigos_programa = []

            for _, row in df_programa.iterrows():
                producto = row["Nombre STD"]
                toneladas = row["Toneladas"]

                if "Código Canal" in df_ddp.columns:
                    codigos_producto = df_ddp[df_ddp["Producto"] == producto]["Código Canal"].dropna().unique()
                    for codigo in codigos_producto:
                        # Los pases falsos no son canales: no se cuentan como cilindro
                        # ni se les pide rendimiento.
                        if _es_pase_falso(codigo):
                            continue
                        codigos_programa.append({
                            "Nombre STD": producto,
                            "Código Canal": codigo,
                            "Toneladas": toneladas
                        })
            
            # Convertir a DataFrame
            df_codigos_programa = pd.DataFrame(codigos_programa)
            
            # Calcular frecuencia si hay datos
            if not df_codigos_programa.empty:
                frecuencia_en_programa = (
                    df_codigos_programa
                    .groupby("Código Canal", dropna=True)
                    .agg(
                        Frecuencia=("Nombre STD", "count"),
                        Toneladas=("Toneladas", "sum")
                    )
                    .reset_index()
                    .sort_values("Toneladas", ascending=False)
                )

                # --- Canales y cilindros requeridos, usando la segunda hoja del Consolidado ---
                try:
                    frecuencia_en_programa = _calcular_canales_cilindros(frecuencia_en_programa, df_rendimiento)
                    sin_rendimiento = frecuencia_en_programa.loc[
                        frecuencia_en_programa["Observación"] == "Sin rendimiento registrado en Consolidado",
                        "Código Canal"
                    ].tolist()
                    if sin_rendimiento:
                        st.warning(
                            f"{len(sin_rendimiento)} código(s) de canal sin rendimiento registrado en el "
                            f"Consolidado — no fue posible estimar canales/cilindros requeridos: "
                            + ", ".join(sin_rendimiento)
                        )
                except Exception as e:
                    logger.error(f"Error calculando canales/cilindros requeridos: {str(e)}")
                    st.warning("No se pudieron calcular los canales/cilindros requeridos (rendimiento).")

                # Mostrar tabla de frecuencias
                st.dataframe(frecuencia_en_programa.set_index("Código Canal"), width="stretch")
                
                # Mostrar métricas de cilindros
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("Cilindros Únicos", len(frecuencia_en_programa))
                with col2:
                    cilindro_mas_usado = frecuencia_en_programa.iloc[0]["Código Canal"] if not frecuencia_en_programa.empty else "N/A"
                    st.metric("Cilindro Más Usado", cilindro_mas_usado)
                with col3:
                    max_frecuencia = frecuencia_en_programa.iloc[0]["Frecuencia"] if not frecuencia_en_programa.empty else 0
                    st.metric("Frecuencia Máxima", max_frecuencia)

            else:
                st.warning("No se encontraron códigos de canal para los productos en el programa.")
                frecuencia_en_programa = pd.DataFrame()

        except Exception as e:
            logger.error(f"Error calculando frecuencia de cilindros: {str(e)}")
            st.error("Error calculando frecuencia de cilindros")
            frecuencia_en_programa = pd.DataFrame()
        
        # ===============================================
        # EXPORTACIÓN
        # ===============================================

        st.markdown("---")
        st.markdown("### Exportar Datos")

        col1, col2 = st.columns(2)

        with col1:
            try:
                buffer_completo = _exportar_maestranza_xlsx(
                    df_resumen, frecuencia_en_programa, st.session_state.df_prog
                )
                timestamp = datetime.now().strftime("%Y%m%d_%H%M")
                st.download_button(
                    label="Descargar Resumen Técnico Completo",
                    data=buffer_completo,
                    file_name=f"Resumen_Maestranza_{timestamp}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    help="Incluye resumen maestranza, frecuencia de cilindros y programa completo"
                )
            except Exception as e:
                logger.error(f"Error creando Excel: {str(e)}")
                st.error("Error generando archivo Excel")
        
    except Exception as e:
        st.error(f"Error generando resumen maestranza: {str(e)}")
        logger.error(f"Error en mostrar_resumen_maestranza: {str(e)}")

def mostrar_analisis_utilaje(df_ddp):
    """Muestra el análisis detallado de utilaje."""
    
    try:
        # Verificar que tenemos los datos necesarios
        if df_ddp.empty:
            st.warning("No hay datos de productos disponibles para análisis de utilaje.")
            return
        
        # Definir componentes de utilaje
        componentes_utilaje = [
            "Caja Guía Entrada",
            "Caja Guía Salida", 
            "Embudo Entrada",
            "Embudo Salida",
            "Código Polín Entrada",
            "Código Polín Salida",
            "Estabilización Entrada", 
            "Estabilización Salida",
            "Rodamiento Entrada",
            "Rodamiento Salida",
            "Semiguía Entrada",
            "Semiguía Salida",
            "Raspador Entrada",
            "Raspador Salida"
        ]
        
        # Verificar qué componentes existen en los datos
        componentes_disponibles = [comp for comp in componentes_utilaje if comp in df_ddp.columns]
        componentes_faltantes = [comp for comp in componentes_utilaje if comp not in df_ddp.columns]
        
        # Mostrar información de disponibilidad
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("Total Productos", len(df_ddp["Producto"].unique()) if "Producto" in df_ddp.columns else 0)
        with col2:
            st.metric("Componentes Disponibles", len(componentes_disponibles))
        with col3:
            st.metric("Componentes Faltantes", len(componentes_faltantes))
        
        if componentes_faltantes:
            with st.expander("Componentes no encontrados en los datos"):
                for comp in componentes_faltantes:
                    st.write(f"• {comp}")
        
        if not componentes_disponibles:
            st.error("No se encontraron componentes de utilaje en los datos.")
            return
        
        # Crear pestañas para diferentes análisis
        sub_tabs = st.tabs([
            "Análisis según Programa",
            "Análisis Individual", 
            "Comparación Manual",
            "Estadísticas Generales"
        ])
        
        # PESTAÑA 1: ANÁLISIS SEGÚN PROGRAMA
        with sub_tabs[0]:
            if "df_prog" in st.session_state:
                mostrar_utilaje_programa(df_ddp, componentes_disponibles)
            else:
                st.info("Por favor carga primero el archivo de programa para ver el análisis de utilaje según la secuencia de producción.")
        
        # PESTAÑA 2: ANÁLISIS INDIVIDUAL
        with sub_tabs[1]:
            st.markdown("### Análisis Individual de Producto")
            
            productos_disponibles = sorted(df_ddp["Producto"].dropna().unique()) if "Producto" in df_ddp.columns else []
            
            if productos_disponibles:
                col_prod, col_filtro = st.columns([2, 1])
                
                with col_prod:
                    producto_seleccionado = st.selectbox(
                        "Selecciona un producto para ver su utilaje:",
                        productos_disponibles,
                        key="producto_utilaje_individual",
                        format_func=etiqueta_producto
                    )
                
                with col_filtro:
                    st.markdown("**Opciones:**")
                    mostrar_solo_definidos = st.checkbox("Solo mostrar componentes definidos", value=True, key="filtro_individual")
                
                if producto_seleccionado:
                    mostrar_utilaje_producto(df_ddp, producto_seleccionado, componentes_disponibles, mostrar_solo_definidos)
        
        # PESTAÑA 3: COMPARACIÓN MANUAL
        with sub_tabs[2]:
            st.markdown("### Comparación de Utilaje entre Productos")
            
            productos_disponibles = sorted(df_ddp["Producto"].dropna().unique()) if "Producto" in df_ddp.columns else []
            
            if len(productos_disponibles) >= 2:
                col_a, col_b, col_opciones = st.columns([2, 2, 1])
                
                with col_a:
                    producto_a_util = st.selectbox(
                        "Producto A:",
                        productos_disponibles,
                        key="producto_a_utilaje_comp",
                        format_func=etiqueta_producto
                    )
                
                with col_b:
                    producto_b_util = st.selectbox(
                        "Producto B:",
                        productos_disponibles,
                        index=1 if len(productos_disponibles) > 1 else 0,
                        key="producto_b_utilaje_comp",
                        format_func=etiqueta_producto
                    )
                
                with col_opciones:
                    st.markdown("**Opciones:**")
                    solo_diferencias = st.checkbox("Solo diferencias", value=True, key="solo_dif_utilaje_comp")
                
                if producto_a_util != producto_b_util:
                    comparar_utilaje_productos(df_ddp, producto_a_util, producto_b_util, componentes_disponibles, solo_diferencias)
                else:
                    st.warning("Selecciona productos diferentes para compararlos.")
            else:
                st.warning("Se necesitan al menos 2 productos para comparar.")
        
        # PESTAÑA 4: ESTADÍSTICAS GENERALES
        with sub_tabs[3]:
            st.markdown("### Análisis General de Utilaje")
            mostrar_estadisticas_utilaje(df_ddp, componentes_disponibles)
        
    except Exception as e:
        st.error(f"Error en análisis de utilaje: {str(e)}")
        logger.error(f"Error en mostrar_analisis_utilaje: {str(e)}")

def mostrar_utilaje_programa(df_ddp, componentes_disponibles):
    """Muestra el análisis de utilaje basado en el programa de producción."""
    
    try:
        df_prog = st.session_state.df_prog.copy()
        
        st.markdown("### Análisis de Utilaje según Secuencia de Producción")
        
        with st.spinner("Analizando necesidades de utilaje según programa..."):
            # Detectar bloques consecutivos del mismo producto
            df_prog["Grupo"] = (df_prog["Nombre STD"] != df_prog["Nombre STD"].shift()).cumsum()
            
            # Verificar que existe la columna PROGR para toneladas
            if "PROGR" not in df_prog.columns:
                st.error("El archivo de programa debe contener la columna 'PROGR' para calcular toneladas")
                return
            
            # Agrupar y sumar toneladas
            df_programa = (
                df_prog.dropna(subset=["PROGR"])
                .groupby(["Grupo", "Nombre STD"], as_index=False)
                .agg({"PROGR": "sum"})
                .rename(columns={"PROGR": "Toneladas"})
            )
            df_programa = df_programa[df_programa["Toneladas"] > 0].reset_index(drop=True)
            df_programa["Toneladas"] = df_programa["Toneladas"].astype(int)
            
            # Análisis de cambios de utilaje en la secuencia
            cambios_utilaje = []
            
            for i in range(len(df_programa) - 1):
                producto_actual = df_programa.loc[i, "Nombre STD"]
                producto_siguiente = df_programa.loc[i + 1, "Nombre STD"]
                
                # Skip si es el mismo producto
                if producto_actual == producto_siguiente:
                    continue
                
                # Obtener datos de utilaje para ambos productos
                datos_actual = df_ddp[df_ddp["Producto"] == producto_actual]
                datos_siguiente = df_ddp[df_ddp["Producto"] == producto_siguiente]
                
                if datos_actual.empty or datos_siguiente.empty:
                    continue
                
                # Analizar cambios en cada componente
                componentes_cambian = []
                for comp in componentes_disponibles:
                    val_actual = datos_actual[comp].dropna().unique()
                    val_siguiente = datos_siguiente[comp].dropna().unique()
                    
                    # Determinar si hay cambio
                    if len(val_actual) > 0 and len(val_siguiente) > 0:
                        if set(val_actual) != set(val_siguiente):
                            componentes_cambian.append(comp)
                
                if componentes_cambian:
                    cambios_utilaje.append({
                        "Secuencia": i + 1,
                        "Producto Origen": producto_actual,
                        "Producto Destino": producto_siguiente,
                        "Componentes que Cambian": len(componentes_cambian),
                        "Detalle Componentes": ", ".join(componentes_cambian[:3]) + ("..." if len(componentes_cambian) > 3 else "")
                    })
        
        # Mostrar métricas generales
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("Total Productos", len(df_programa))
        with col2:
            st.metric("Cambios de Producto", len(cambios_utilaje))
        with col3:
            cambios_con_utilaje = len([c for c in cambios_utilaje if c["Componentes que Cambian"] > 0])
            st.metric("Cambios con Utilaje", cambios_con_utilaje)
        with col4:
            total_toneladas = df_programa["Toneladas"].sum()
            st.metric("Total Toneladas", f"{total_toneladas:,.0f}")
        
        # Mostrar tabla de cambios de utilaje
        if cambios_utilaje:
            st.markdown("---")
            st.markdown("#### Cambios de Utilaje en la Secuencia")
            
            df_cambios = pd.DataFrame(cambios_utilaje)
            
            # Aplicar color según cantidad de cambios
            def colorear_cambios(row):
                cambios = row["Componentes que Cambian"]
                if cambios >= 10:
                    return [f'background-color: #ffcdd2; color: {TEXTO_PASTEL}'] * len(row)  # Rojo claro
                elif cambios >= 5:
                    return [f'background-color: #fff9c4; color: {TEXTO_PASTEL}'] * len(row)  # Amarillo claro
                else:
                    return [f'background-color: #c8e6c9; color: {TEXTO_PASTEL}'] * len(row)  # Verde claro
            
            st.dataframe(
                df_cambios.style.apply(colorear_cambios, axis=1),
                width="stretch",
                hide_index=True
            )
            
            # Leyenda de colores
            with st.expander("Leyenda de colores"):
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.markdown(_muestra_color("#c8e6c9", "<b>1-4</b> componentes cambian"),
                                unsafe_allow_html=True)
                with col2:
                    st.markdown(_muestra_color("#fff9c4", "<b>5-9</b> componentes cambian"),
                                unsafe_allow_html=True)
                with col3:
                    st.markdown(_muestra_color("#ffcdd2", "<b>10+</b> componentes cambian"),
                                unsafe_allow_html=True)
        
        # Análisis de frecuencia de utilajes en el programa
        st.markdown("---")
        st.markdown("#### Frecuencia de Utilajes en Programa")
        
        # Recopilar todos los utilajes usados en el programa
        utilajes_programa = {}
        
        for componente in componentes_disponibles:
            utilajes_programa[componente] = []
            
            for _, row in df_programa.iterrows():
                producto = row["Nombre STD"]
                toneladas = row["Toneladas"]
                
                # Obtener valores de utilaje para este producto
                datos_producto = df_ddp[df_ddp["Producto"] == producto]
                if not datos_producto.empty:
                    valores = datos_producto[componente].dropna().unique()
                    for valor in valores:
                        utilajes_programa[componente].append({
                            "Valor": valor,
                            "Producto": producto,
                            "Toneladas": toneladas
                        })
        
        # Selector de componente para análisis detallado
        componente_analizar = st.selectbox(
            "Selecciona componente para ver frecuencia:",
            componentes_disponibles,
            key="comp_frecuencia_programa"
        )
        
        if componente_analizar and utilajes_programa[componente_analizar]:
            df_comp = pd.DataFrame(utilajes_programa[componente_analizar])
            
            # Agrupar por valor y sumar toneladas
            frecuencia_comp = (
                df_comp
                .groupby("Valor")
                .agg(
                    Frecuencia=("Producto", "count"),
                    Toneladas_Total=("Toneladas", "sum"),
                    Productos=("Producto", lambda x: ", ".join(x.unique()[:3]) + ("..." if len(x.unique()) > 3 else ""))
                )
                .reset_index()
                .sort_values("Toneladas_Total", ascending=False)
            )
            
            col1, col2 = st.columns([2, 1])
            
            with col1:
                st.dataframe(frecuencia_comp, width="stretch", hide_index=True)
            
            with col2:
                # Métricas del componente
                st.metric("Valores Únicos", len(frecuencia_comp))
                if not frecuencia_comp.empty:
                    valor_mas_usado = frecuencia_comp.iloc[0]["Valor"]
                    st.metric("Más Usado", valor_mas_usado)
        
        # Resumen de necesidades de utilaje
        st.markdown("---")
        st.markdown("#### Resumen de Necesidades de Utilaje")
        
        # Crear tabla resumen con todos los productos y sus utilajes
        resumen_utilaje = []
        
        for _, row in df_programa.iterrows():
            producto = row["Nombre STD"]
            toneladas = row["Toneladas"]
            
            datos_producto = df_ddp[df_ddp["Producto"] == producto]
            if not datos_producto.empty:
                fila_resumen = {
                    "Producto": producto,
                    "Toneladas": toneladas
                }
                
                # Agregar valores de cada componente
                for comp in componentes_disponibles[:6]:  # Mostrar solo los primeros 6 para no hacer la tabla muy ancha
                    valores = datos_producto[comp].dropna().unique()
                    fila_resumen[comp] = ", ".join([str(v) for v in valores]) if len(valores) > 0 else "-"
                
                resumen_utilaje.append(fila_resumen)
        
        if resumen_utilaje:
            df_resumen_utilaje = pd.DataFrame(resumen_utilaje)
            st.dataframe(df_resumen_utilaje, width="stretch", hide_index=True)
        
        # Botón de exportación
        st.markdown("---")
        col1, col2, col3 = st.columns([1, 1, 2])
        
        with col1:
            if st.button("Exportar Análisis de Utilaje", key="export_utilaje_programa"):
                exportar_utilaje_programa(df_programa, df_ddp, componentes_disponibles, cambios_utilaje)
        
    except Exception as e:
        st.error(f"Error en análisis de utilaje según programa: {str(e)}")
        logger.error(f"Error en mostrar_utilaje_programa: {str(e)}")

def exportar_utilaje_programa(df_programa, df_ddp, componentes_disponibles, cambios_utilaje):
    """Exporta el análisis de utilaje según programa."""
    
    try:
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine="xlsxwriter") as writer:
            # Hoja 1: Resumen del programa con utilajes principales
            resumen_export = []
            for _, row in df_programa.iterrows():
                producto = row["Nombre STD"]
                toneladas = row["Toneladas"]
                
                datos_producto = df_ddp[df_ddp["Producto"] == producto]
                if not datos_producto.empty:
                    fila = {"Producto": producto, "Toneladas": toneladas}
                    
                    for comp in componentes_disponibles:
                        valores = datos_producto[comp].dropna().unique()
                        fila[comp] = ", ".join([str(v) for v in valores]) if len(valores) > 0 else ""
                    
                    resumen_export.append(fila)
            
            if resumen_export:
                df_resumen = pd.DataFrame(resumen_export)
                df_resumen.to_excel(writer, sheet_name="Resumen_Utilaje_Programa", index=False)
            
            # Hoja 2: Cambios de utilaje
            if cambios_utilaje:
                df_cambios = pd.DataFrame(cambios_utilaje)
                df_cambios.to_excel(writer, sheet_name="Cambios_Utilaje", index=False)
            
            # Hoja 3: Frecuencia de cada componente
            for i, comp in enumerate(componentes_disponibles[:10]):  # Limitar a 10 componentes
                frecuencias = {}
                
                for _, row in df_programa.iterrows():
                    producto = row["Nombre STD"]
                    toneladas = row["Toneladas"]
                    
                    datos_producto = df_ddp[df_ddp["Producto"] == producto]
                    if not datos_producto.empty:
                        valores = datos_producto[comp].dropna().unique()
                        for valor in valores:
                            if valor not in frecuencias:
                                frecuencias[valor] = {"Frecuencia": 0, "Toneladas": 0, "Productos": []}
                            frecuencias[valor]["Frecuencia"] += 1
                            frecuencias[valor]["Toneladas"] += toneladas
                            frecuencias[valor]["Productos"].append(producto)
                
                if frecuencias:
                    df_freq = pd.DataFrame([
                        {
                            "Valor": k,
                            "Frecuencia": v["Frecuencia"],
                            "Toneladas": v["Toneladas"],
                            "Productos": ", ".join(v["Productos"][:5]) + ("..." if len(v["Productos"]) > 5 else "")
                        }
                        for k, v in frecuencias.items()
                    ]).sort_values("Toneladas", ascending=False)
                    
                    sheet_name = f"Frec_{comp[:20]}" if i < 9 else "Frec_Otros"  # Limitar nombre de hoja
                    df_freq.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Formatear hojas
            workbook = writer.book
            header_format = workbook.add_format({
                'bold': True,
                'bg_color': '#4CAF50',
                'font_color': 'white',
                'border': 1
            })
            
            for sheet_name in writer.sheets:
                worksheet = writer.sheets[sheet_name]
                worksheet.set_row(0, 20, header_format)
                worksheet.autofit()
        
        buffer.seek(0)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M")
        filename = f"Analisis_Utilaje_Programa_{timestamp}.xlsx"
        
        st.download_button(
            label="Descargar Análisis Completo",
            data=buffer,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            help="Incluye resumen de utilaje por programa, cambios y frecuencias"
        )
        
        st.success("Análisis exportado exitosamente")
        
    except Exception as e:
        st.error(f"Error exportando análisis: {str(e)}")
        logger.error(f"Error en exportar_utilaje_programa: {str(e)}")

def mostrar_utilaje_producto(df_ddp, producto, componentes_disponibles, mostrar_solo_definidos=True):
    """Muestra el detalle de utilaje para un producto específico."""
    
    try:
        # Filtrar datos del producto
        datos_producto = df_ddp[df_ddp["Producto"] == producto]
        
        if datos_producto.empty:
            st.warning(f"No se encontraron datos para el producto {producto}")
            return
        
        # Crear tabla de utilaje
        utilaje_data = []
        
        for componente in componentes_disponibles:
            valores_componente = datos_producto[componente].dropna().unique()
            
            # Si solo mostrar definidos y no hay valores, saltar
            if mostrar_solo_definidos and (len(valores_componente) == 0 or (len(valores_componente) == 1 and pd.isna(valores_componente[0]))):
                continue
            
            # Procesar valores
            if len(valores_componente) == 0:
                valor_mostrar = "No definido"
            elif len(valores_componente) == 1:
                valor_mostrar = str(valores_componente[0]) if not pd.isna(valores_componente[0]) else "No definido"
            else:
                valor_mostrar = ", ".join([str(v) for v in valores_componente if not pd.isna(v)])
            
            utilaje_data.append({
                "Componente": componente,
                "Valor": valor_mostrar,
                "Múltiples Valores": "Sí" if len(valores_componente) > 1 else "No"
            })
        
        if utilaje_data:
            df_utilaje = pd.DataFrame(utilaje_data)
            
            # Aplicar estilo condicional
            def resaltar_multiples(row):
                if row["Múltiples Valores"] == "Sí":
                    return [f'background-color: #fff3cd; color: {TEXTO_PASTEL}'] * len(row)
                elif row["Valor"] == "No definido":
                    return [f'background-color: #f8d7da; color: {TEXTO_PASTEL}'] * len(row)
                else:
                    return [f'background-color: #d1edff; color: {TEXTO_PASTEL}'] * len(row)
            
            st.dataframe(
                df_utilaje.style.apply(resaltar_multiples, axis=1),
                width="stretch",
                hide_index=True
            )
            
            # Mostrar leyenda de colores
            with st.expander("Leyenda de colores"):
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.markdown(_muestra_color("#d1edff", "Valor único definido"),
                                unsafe_allow_html=True)
                with col2:
                    st.markdown(_muestra_color("#fff3cd", "Múltiples valores"),
                                unsafe_allow_html=True)
                with col3:
                    st.markdown(_muestra_color("#f8d7da", "No definido"),
                                unsafe_allow_html=True)
        else:
            st.info("No hay componentes de utilaje definidos para este producto (o todos están ocultos por el filtro).")
            
    except Exception as e:
        st.error(f"Error mostrando utilaje del producto: {str(e)}")
        logger.error(f"Error en mostrar_utilaje_producto: {str(e)}")

def comparar_utilaje_productos(df_ddp, producto_a, producto_b, componentes_disponibles, solo_diferencias=True):
    """Compara el utilaje entre dos productos."""
    
    try:
        # Obtener datos de ambos productos
        datos_a = df_ddp[df_ddp["Producto"] == producto_a]
        datos_b = df_ddp[df_ddp["Producto"] == producto_b]
        
        if datos_a.empty or datos_b.empty:
            st.warning("No se encontraron datos para uno o ambos productos.")
            return
        
        # Crear comparación
        comparacion_data = []
        
        for componente in componentes_disponibles:
            # Obtener valores únicos de cada producto
            valores_a = datos_a[componente].dropna().unique()
            valores_b = datos_b[componente].dropna().unique()
            
            # Procesar valores para mostrar
            valor_a = ", ".join([str(v) for v in valores_a]) if len(valores_a) > 0 else "No definido"
            valor_b = ", ".join([str(v) for v in valores_b]) if len(valores_b) > 0 else "No definido"
            
            # Determinar si hay diferencia
            diferentes = set(valores_a) != set(valores_b)
            
            # Si solo mostrar diferencias y son iguales, saltar
            if solo_diferencias and not diferentes:
                continue
            
            comparacion_data.append({
                "Componente": componente,
                f"Producto A ({producto_a})": valor_a,
                f"Producto B ({producto_b})": valor_b,
                "¿Diferente?": "Sí" if diferentes else "No"
            })
        
        if comparacion_data:
            df_comparacion = pd.DataFrame(comparacion_data)
            
            # Mostrar métricas de comparación
            total_componentes = len(df_comparacion)
            componentes_diferentes = len(df_comparacion[df_comparacion["¿Diferente?"] == "Sí"])
            porcentaje_diferencias = (componentes_diferentes / total_componentes * 100) if total_componentes > 0 else 0
            
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("Componentes Analizados", total_componentes)
            with col2:
                st.metric("Componentes Diferentes", componentes_diferentes)
            with col3:
                st.metric("% Diferencias", f"{porcentaje_diferencias:.1f}%")
            
            # Aplicar estilo a la tabla
            def resaltar_diferencias(row):
                if row["¿Diferente?"] == "Sí":
                    return [f'background-color: #ffebee; color: {TEXTO_PASTEL}'] * len(row)
                else:
                    return [f'background-color: #f1f8e9; color: {TEXTO_PASTEL}'] * len(row)
            
            st.dataframe(
                df_comparacion.style.apply(resaltar_diferencias, axis=1),
                width="stretch",
                hide_index=True
            )
            
        else:
            if solo_diferencias:
                st.success("**¡No hay diferencias en el utilaje entre estos productos!**")
            else:
                st.info("No se encontraron componentes de utilaje para comparar.")
                
    except Exception as e:
        st.error(f"Error comparando utilaje: {str(e)}")
        logger.error(f"Error en comparar_utilaje_productos: {str(e)}")

def mostrar_estadisticas_utilaje(df_ddp, componentes_disponibles):
    """Muestra estadísticas generales del utilaje."""
    
    try:
        # Análisis de frecuencias por componente
        st.markdown("####Frecuencia de Valores por Componente")
        
        # Selector de componente para análisis detallado
        componente_analisis = st.selectbox(
            "Selecciona componente para análisis detallado:",
            componentes_disponibles,
            key="componente_analisis"
        )
        
        if componente_analisis:
            col_analisis, col_exportar = st.columns([3, 1])
            
            with col_analisis:
                # Obtener frecuencias del componente seleccionado
                valores_componente = df_ddp[componente_analisis].dropna()
                
                if not valores_componente.empty:
                    frecuencias = valores_componente.value_counts().reset_index()
                    frecuencias.columns = ["Valor", "Frecuencia"]
                    frecuencias["Porcentaje"] = (frecuencias["Frecuencia"] / frecuencias["Frecuencia"].sum() * 100).round(1)
                    
                    st.dataframe(frecuencias, width="stretch", hide_index=True)
                    
                    # Mostrar métricas del componente
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.metric("Valores Únicos", len(frecuencias))
                    with col2:
                        st.metric("Productos con Valor", len(valores_componente))
                    with col3:
                        valor_mas_comun = frecuencias.iloc[0]["Valor"] if not frecuencias.empty else "N/A"
                        st.metric("Valor Más Común", valor_mas_comun)
                else:
                    st.info(f"No hay valores definidos para {componente_analisis}")
            
            with col_exportar:
                # Botón para exportar análisis completo
                if st.button("Exportar Análisis"):
                    exportar_analisis_utilaje(df_ddp, componentes_disponibles)
        
        # Resumen general de todos los componentes
        st.markdown("---")
        st.markdown("#### Resumen General de Componentes")
        
        resumen_general = []
        for componente in componentes_disponibles:
            valores = df_ddp[componente].dropna()
            valores_unicos = len(valores.unique()) if not valores.empty else 0
            productos_con_valor = len(valores) if not valores.empty else 0
            total_productos = len(df_ddp)
            cobertura = (productos_con_valor / total_productos * 100) if total_productos > 0 else 0
            
            resumen_general.append({
                "Componente": componente,
                "Valores Únicos": valores_unicos,
                "Productos con Valor": productos_con_valor,
                "Cobertura (%)": f"{cobertura:.1f}%"
            })
        
        df_resumen = pd.DataFrame(resumen_general)
        
        # Colorear por cobertura
        def colorear_cobertura(row):
            cobertura = float(row["Cobertura (%)"].replace("%", ""))
            if cobertura >= 80:
                return [f'background-color: #d1edff; color: {TEXTO_PASTEL}'] * len(row)
            elif cobertura >= 50:
                return [f'background-color: #fff3cd; color: {TEXTO_PASTEL}'] * len(row)
            else:
                return [f'background-color: #f8d7da; color: {TEXTO_PASTEL}'] * len(row)
        
        st.dataframe(
            df_resumen.style.apply(colorear_cobertura, axis=1),
            width="stretch",
            hide_index=True
        )
        
    except Exception as e:
        st.error(f"Error en estadísticas de utilaje: {str(e)}")
        logger.error(f"Error en mostrar_estadisticas_utilaje: {str(e)}")

def exportar_analisis_utilaje(df_ddp, componentes_disponibles):
    """Exporta el análisis completo de utilaje."""
    
    try:
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine="xlsxwriter") as writer:
            # Hoja 1: Resumen por componente
            resumen_componentes = []
            for componente in componentes_disponibles:
                valores = df_ddp[componente].dropna()
                frecuencias = valores.value_counts()
                
                for valor, freq in frecuencias.items():
                    resumen_componentes.append({
                        "Componente": componente,
                        "Valor": valor,
                        "Frecuencia": freq
                    })
            
            if resumen_componentes:
                df_resumen_comp = pd.DataFrame(resumen_componentes)
                df_resumen_comp.to_excel(writer, sheet_name="Resumen_Componentes", index=False)
            
            # Hoja 2: Matriz completa de utilaje
            columnas_utilaje = ["Producto"] + componentes_disponibles
            df_utilaje_completo = df_ddp[columnas_utilaje]
            df_utilaje_completo.to_excel(writer, sheet_name="Utilaje_Completo", index=False)
            
            # Hoja 3: Estadísticas generales
            estadisticas = []
            for componente in componentes_disponibles:
                valores = df_ddp[componente].dropna()
                estadisticas.append({
                    "Componente": componente,
                    "Valores_Unicos": len(valores.unique()) if not valores.empty else 0,
                    "Productos_con_Valor": len(valores),
                    "Cobertura_Porcentaje": len(valores) / len(df_ddp) * 100 if len(df_ddp) > 0 else 0
                })
            
            df_estadisticas = pd.DataFrame(estadisticas)
            df_estadisticas.to_excel(writer, sheet_name="Estadisticas", index=False)
        
        buffer.seek(0)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M")
        filename = f"Analisis_Utilaje_{timestamp}.xlsx"
        
        st.download_button(
            label="Descargar Análisis Completo",
            data=buffer,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            help="Incluye resumen por componentes, matriz completa y estadísticas"
        )
        
        st.success("Análisis exportado exitosamente")
        
    except Exception as e:
        st.error(f"Error exportando análisis: {str(e)}")
        logger.error(f"Error en exportar_analisis_utilaje: {str(e)}")

if __name__ == "__main__":
    main()